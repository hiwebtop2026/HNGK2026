# -*- coding: utf-8 -*-
"""
Excel样式优化工具
直接修改printPage (1)_分析版.xlsx文件的排版和配色方案
实现现代风格的UI设计，提高数据易读性
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, GradientFill
from openpyxl.utils import get_column_letter
import os

INPUT_FILE = r'C:\Users\lhp\Desktop\printPage (1)_分析版_backup.xlsx'
OUTPUT_FILE = r'C:\Users\lhp\Desktop\printPage (1)_分析版.xlsx'

COLORS = {
    'bg_page': 'F8FAFC',
    'header': '1E293B',
    'header_bg': '334155',
    'card_bg': 'FFFFFF',
    'card_border': 'E2E8F0',
    'text_primary': '1E293B',
    'text_secondary': '64748B',
    'prob_high': '10B981',
    'prob_high_bg': 'D1FAE5',
    'prob_medium_high': '3B82F6',
    'prob_medium_high_bg': 'DBEAFE',
    'prob_medium': 'F59E0B',
    'prob_medium_bg': 'FEF3C7',
    'prob_low': 'EF4444',
    'prob_low_bg': 'FEE2E2',
    'prob_very_low': 'DC2626',
    'prob_very_low_bg': 'FECACA',
    'school_985': '7C3AED',
    'school_double': '3B82F6',
    'school_normal': '10B981',
}

FONT_HEADER = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
FONT_TITLE = Font(name='微软雅黑', size=14, bold=True, color=COLORS['header'])
FONT_NORMAL = Font(name='微软雅黑', size=11, color=COLORS['text_primary'])
FONT_SECONDARY = Font(name='微软雅黑', size=10, color=COLORS['text_secondary'])
FONT_BOLD = Font(name='微软雅黑', size=11, bold=True, color=COLORS['text_primary'])

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center', wrap_text=True)

BORDER_THIN = Border(
    left=Side(style='thin', color=COLORS['card_border']),
    right=Side(style='thin', color=COLORS['card_border']),
    top=Side(style='thin', color=COLORS['card_border']),
    bottom=Side(style='thin', color=COLORS['card_border'])
)

BORDER_THICK = Border(
    left=Side(style='medium', color='CBD5E1'),
    right=Side(style='medium', color='CBD5E1'),
    top=Side(style='medium', color='CBD5E1'),
    bottom=Side(style='medium', color='CBD5E1')
)

FILL_HEADER = PatternFill(start_color=COLORS['header_bg'], end_color=COLORS['header_bg'], fill_type='solid')
FILL_CARD = PatternFill(start_color=COLORS['card_bg'], end_color=COLORS['card_bg'], fill_type='solid')

PROB_COLOR_MAP = {
    '高': {'bg': COLORS['prob_high_bg'], 'font': COLORS['prob_high']},
    '偏高': {'bg': COLORS['prob_medium_high_bg'], 'font': COLORS['prob_medium_high']},
    '中': {'bg': COLORS['prob_medium_bg'], 'font': COLORS['prob_medium']},
    '偏低': {'bg': COLORS['prob_low_bg'], 'font': COLORS['prob_low']},
    '低': {'bg': COLORS['prob_very_low_bg'], 'font': COLORS['prob_very_low']},
}


def get_probability_level(prob_str):
    try:
        prob = int(prob_str.replace('%', ''))
        if prob >= 85:
            return '高'
        elif prob >= 70:
            return '偏高'
        elif prob >= 50:
            return '中'
        elif prob >= 30:
            return '偏低'
        else:
            return '低'
    except:
        return None


def apply_school_type_style(ws, row, col, school_name):
    cell = ws.cell(row=row, column=col)
    if '985' in school_name or school_name in ['清华大学', '北京大学', '中国人民大学', '北京师范大学', 
                                               '北京航空航天大学', '北京理工大学', '中国农业大学', 
                                               '中央民族大学', '吉林大学', '东北大学']:
        cell.font = Font(name='微软雅黑', size=11, bold=True, color=COLORS['school_985'])
    elif '双一流' in school_name or school_name in ['南京理工大学', '河海大学', '苏州大学', '合肥工业大学',
                                                   '武汉理工大学', '中国矿业大学', '南京邮电大学', '江南大学',
                                                   '西南交通大学', '安徽大学', '西南财经大学', '华东理工大学',
                                                   '华南师范大学', '东华大学', '中国地质大学(武汉)', 
                                                   '长安大学', '北京林业大学', '暨南大学']:
        cell.font = Font(name='微软雅黑', size=11, bold=True, color=COLORS['school_double'])
    else:
        cell.font = Font(name='微软雅黑', size=11, color=COLORS['school_normal'])


def optimize_excel_style():
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active
    ws.title = '志愿分析'
    
    max_row = ws.max_row
    max_col = ws.max_column
    
    for col in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    ws.column_dimensions['C'].width = 80
    ws.column_dimensions['G'].width = 60
    
    header_row = 1
    headers = ['序号', '类型', '内容', '是否服从', '录取概率', '概率等级', '分析']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THICK
    
    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDER_THIN
            
            if col == 1:
                cell.alignment = ALIGN_CENTER
                cell.font = FONT_BOLD if cell.value else FONT_NORMAL
            elif col == 2:
                cell.alignment = ALIGN_CENTER
                if cell.value == '院校专业组':
                    cell.fill = PatternFill(start_color='E0F2FE', end_color='E0F2FE', fill_type='solid')
                    cell.font = Font(name='微软雅黑', size=11, bold=True, color='0369A1')
                elif cell.value == '专业调剂':
                    cell.fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
                    cell.font = Font(name='微软雅黑', size=10, color='D97706')
                elif cell.value == '专业':
                    cell.fill = PatternFill(start_color='ECFDF5', end_color='ECFDF5', fill_type='solid')
                    cell.font = Font(name='微软雅黑', size=11, bold=True, color='059669')
            elif col == 3:
                cell.alignment = ALIGN_LEFT
                cell.font = FONT_NORMAL
            elif col == 4:
                cell.alignment = ALIGN_CENTER
                if cell.value == '服从':
                    cell.fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
                    cell.font = Font(name='微软雅黑', size=11, bold=True, color='059669')
                elif cell.value == '不服从':
                    cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
                    cell.font = Font(name='微软雅黑', size=11, bold=True, color='DC2626')
                else:
                    cell.font = FONT_SECONDARY
            elif col == 5:
                cell.alignment = ALIGN_CENTER
                prob_level = get_probability_level(str(cell.value))
                if prob_level and prob_level in PROB_COLOR_MAP:
                    style = PROB_COLOR_MAP[prob_level]
                    cell.fill = PatternFill(start_color=style['bg'], end_color=style['bg'], fill_type='solid')
                    cell.font = Font(name='微软雅黑', size=11, bold=True, color=style['font'])
                else:
                    cell.font = FONT_NORMAL
            elif col == 6:
                cell.alignment = ALIGN_CENTER
                level = str(cell.value).strip()
                if level in PROB_COLOR_MAP:
                    style = PROB_COLOR_MAP[level]
                    cell.fill = PatternFill(start_color=style['bg'], end_color=style['bg'], fill_type='solid')
                    cell.font = Font(name='微软雅黑', size=11, bold=True, color=style['font'])
                else:
                    cell.font = FONT_NORMAL
            elif col == 7:
                cell.alignment = ALIGN_LEFT
                cell.font = FONT_SECONDARY
    
    for row in range(2, max_row + 1):
        if ws.cell(row=row, column=2).value == '院校专业组':
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).border = Border(
                    left=Side(style='medium', color='94A3B8'),
                    right=Side(style='medium', color='94A3B8'),
                    top=Side(style='medium', color='94A3B8'),
                    bottom=Side(style='thin', color='E2E8F0')
                )
        
        if ws.cell(row=row, column=2).value == '专业':
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).border = Border(
                    left=Side(style='medium', color='94A3B8'),
                    right=Side(style='medium', color='94A3B8'),
                    top=Side(style='thin', color='E2E8F0'),
                    bottom=Side(style='medium', color='94A3B8')
                )
    
    ws.sheet_properties.tabColor = COLORS['header_bg']
    
    wb.save(OUTPUT_FILE)
    print(f'✅ Excel样式优化完成！')
    print(f'📁 文件已保存到: {OUTPUT_FILE}')


def add_summary_sheet(wb):
    ws = wb.create_sheet(title='分析汇总')
    
    summary_data = [
        ['考生信息', '', '', '', '', '', ''],
        ['省份', '海南', '', '', '', '', ''],
        ['选科', '物化史', '', '', '', '', ''],
        ['分数', '696', '', '', '', '', ''],
        ['位次', '2015', '', '', '', '', ''],
        ['', '', '', '', '', '', ''],
        ['志愿统计', '', '', '', '', '', ''],
        ['志愿总数', '30', '', '', '', '', ''],
        ['院校数量', '30', '', '', '', '', ''],
        ['专业总数', '172', '', '', '', '', ''],
        ['平均每志愿专业数', '5.7', '', '', '', '', ''],
        ['', '', '', '', '', '', ''],
        ['院校类型分布', '', '', '', '', '', ''],
        ['985', '4', '13.3%', '', '', '', ''],
        ['双一流', '19', '63.3%', '', '', '', ''],
        ['普通本科', '7', '23.3%', '', '', '', ''],
        ['', '', '', '', '', '', ''],
        ['录取概率分布', '', '', '', '', '', ''],
        ['高概率(>85%)', '17', '56.7%', '', '', '', ''],
        ['偏高概率(70-85%)', '4', '13.3%', '', '', '', ''],
        ['中等概率(50-70%)', '3', '10.0%', '', '', '', ''],
        ['偏低概率(30-50%)', '3', '10.0%', '', '', '', ''],
        ['低概率(<30%)', '3', '10.0%', '', '', '', ''],
    ]
    
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    title_cells = [1, 7, 13, 19]
    for row in title_cells:
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.fill = FILL_HEADER
            cell.font = FONT_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_THICK
    
    for row in range(2, len(summary_data) + 1):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDER_THIN
            cell.alignment = ALIGN_CENTER if col <= 3 else ALIGN_LEFT
            if col == 1:
                cell.font = FONT_BOLD
            elif col == 2:
                cell.font = Font(name='微软雅黑', size=11, bold=True, color=COLORS['header'])
            elif col == 3:
                cell.font = Font(name='微软雅黑', size=10, color=COLORS['text_secondary'])
            else:
                cell.font = FONT_NORMAL
    
    ws.sheet_properties.tabColor = COLORS['prob_high']
    
    return wb


def main():
    if os.path.exists(INPUT_FILE):
        print(f'🔧 正在优化Excel样式...')
        wb = openpyxl.load_workbook(INPUT_FILE)
        optimize_excel_style()
        
        wb = openpyxl.load_workbook(OUTPUT_FILE)
        wb = add_summary_sheet(wb)
        wb.save(OUTPUT_FILE)
        
        print(f'\n🎉 Excel样式优化完成！')
        print(f'📁 文件位置: {OUTPUT_FILE}')
        print(f'📊 包含2个工作表: "志愿分析" 和 "分析汇总"')
    else:
        print(f'❌ 文件不存在: {INPUT_FILE}')


if __name__ == "__main__":
    main()