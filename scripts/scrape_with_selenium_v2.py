# -*- coding: utf-8 -*-
"""
夸克高考专业分数线自动抓取工具 v2
使用Selenium + Edge浏览器自动化抓取数据
修复：Excel读取、等待时间、数据提取
"""

import json
import time
import os
import re
from datetime import datetime

from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.microsoft import EdgeChromiumDriverManager

import openpyxl

EXCEL_FILE = r'C:\Users\lhp\Desktop\2023-2025年海南高考本科投档分数线.xlsx'
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data')

YEARS = [2025, 2024, 2023]
PROVINCE = '海南'
BATCH = '本科批'


def ensure_dir(dir_path):
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)


def read_schools_from_excel():
    """从Excel读取院校列表 - 修复版"""
    print('正在读取Excel文件获取院校列表...')
    schools = set()

    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)

        for year in ['2023', '2024', '2025']:
            if year not in wb.sheetnames:
                continue

            ws = wb[year]
            print(f'  读取{year}年...')

            name_col_idx = None
            code_col_idx = None

            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
                for col_idx, cell in enumerate(row):
                    if cell and isinstance(cell, str):
                        cell_str = cell.strip()
                        if ('院校' in cell_str and '名称' in cell_str) or cell_str == '院校专业组名称':
                            name_col_idx = col_idx
                        if ('代码' in cell_str and '院校' in cell_str) or cell_str == '院校专业组代码':
                            code_col_idx = col_idx

                if name_col_idx is not None:
                    print(f'    表头行: {row_idx}, 名称列: {name_col_idx}')
                    break

            if name_col_idx is None:
                print(f'    未找到名称列，跳过')
                continue

            count = 0
            for row in ws.iter_rows(min_row=row_idx + 1, values_only=True):
                if len(row) <= name_col_idx:
                    continue

                cell_val = row[name_col_idx]
                if not cell_val or not isinstance(cell_val, str):
                    continue

                cell_val = cell_val.strip()
                if len(cell_val) < 2:
                    continue

                if '科目' in cell_val and '要求' in cell_val:
                    continue
                if '说明' in cell_val or '注' == cell_val:
                    continue
                if '代码' in cell_val and '名称' in cell_val:
                    continue

                school_name = re.sub(r'\([^)]*\)', '', cell_val).strip()
                school_name = re.sub(r'\d+$', '', school_name).strip()

                if len(school_name) >= 2 and not any(c.isdigit() for c in school_name[:2]):
                    schools.add(school_name)
                    count += 1

            print(f'    提取到 {count} 个院校名称')

        wb.close()
    except Exception as e:
        print(f'读取Excel失败: {e}')
        import traceback
        traceback.print_exc()

    result = sorted(list(schools))
    print(f'共提取 {len(result)} 个唯一院校')

    if len(result) > 0:
        print('前10个院校:')
        for s in result[:10]:
            print(f'  - {s}')

    return result


def setup_driver():
    """设置Edge浏览器"""
    print('正在启动Edge浏览器...')

    options = Options()
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_argument('--start-maximized')
    options.add_argument('--disable-infobars')
    options.add_argument('--disable-gpu')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--window-size=1920,1080')
    options.add_experimental_option('excludeSwitches', ['enable-automation'])
    options.add_experimental_option('useAutomationExtension', False)
    options.page_load_timeout = 60000

    try:
        driver = webdriver.Edge(options=options)
    except Exception as e:
        print(f'默认启动失败，下载驱动: {e}')
        service = Service(EdgeChromiumDriverManager().install())
        driver = webdriver.Edge(service=service, options=options)

    driver.set_page_load_timeout(60)

    driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
        'source': '''
            Object.defineProperty(navigator, 'webdriver', {
                get: () => undefined
            });
            window.chrome = { runtime: {} };
            Object.defineProperty(navigator, 'plugins', {
                get: () => [1, 2, 3, 4, 5]
            });
        '''
    })

    return driver


def build_school_url(school_name, year):
    """构建院校页面URL"""
    params = json.dumps({
        'province': PROVINCE,
        'year': str(year),
        'batch': BATCH,
        'genre': '综合'
    }, ensure_ascii=False)

    base_url = 'https://vt.quark.cn/blm/gaokao-college-794/tab'
    url = (f'{base_url}?app=fen_shu_xian'
           f'&university_name={school_name}'
           f'&q={school_name}'
           f'&device=pc'
           f'&by=tuijian'
           f'&by2=general_entity_college'
           f'&params={params}'
           f'&type=luqu')

    return url


def wait_for_page_load(driver, timeout=20):
    """等待页面加载完成"""
    try:
        WebDriverWait(driver, timeout).until(
            lambda d: d.execute_script('return document.readyState') == 'complete'
        )
    except Exception:
        pass

    time.sleep(5)


def extract_data_via_js(driver, school_name, year):
    """通过JavaScript提取数据 - 增强版"""
    script = '''
    function extractMajorScores(schoolName, year, province, batch) {
        const results = [];

        // 方法1: 从表格提取
        const tables = document.querySelectorAll('table');
        tables.forEach(table => {
            const rows = table.querySelectorAll('tr');
            if (rows.length < 2) return;

            const headers = [];
            const firstCells = rows[0].querySelectorAll('th, td');
            firstCells.forEach(cell => {
                headers.push((cell.textContent || '').trim());
            });

            for (let i = 1; i < rows.length; i++) {
                const cells = rows[i].querySelectorAll('td');
                if (cells.length < 2) continue;

                const record = {
                    school_name: schoolName,
                    province: province,
                    year: year,
                    batch: batch,
                    source: '夸克高考',
                };

                cells.forEach((cell, idx) => {
                    if (idx >= headers.length) return;
                    const header = headers[idx];
                    const value = (cell.textContent || '').trim();

                    if (/专业名称|专业名|专业(?!组|代码)/.test(header)) {
                        record.major_name = value;
                    } else if (/专业组|组名/.test(header)) {
                        record.major_group = value;
                    } else if (/最低分|投档分|分数线/.test(header)) {
                        const num = parseInt(value);
                        if (!isNaN(num)) record.min_score = num;
                    } else if (/最低位次|最低排名|位次/.test(header)) {
                        const num = parseInt(value.replace(/[^\\d]/g, ''));
                        if (!isNaN(num)) record.min_rank = num;
                    } else if (/平均分|均分/.test(header)) {
                        const num = parseInt(value);
                        if (!isNaN(num)) record.avg_score = num;
                    } else if (/科目要求|选科要求|科目/.test(header)) {
                        record.subject_requirement = value;
                    } else if (/招生人数|人数|计划/.test(header)) {
                        const num = parseInt(value);
                        if (!isNaN(num)) record.person_count = num;
                    } else if (/批次/.test(header)) {
                        record.batch = value;
                    }
                });

                if (record.major_name && (record.min_score || record.min_rank)) {
                    results.push(record);
                }
            }
        });

        // 方法2: 从列表项提取
        if (results.length === 0) {
            const items = document.querySelectorAll(
                '[class*="major-item"], [class*="score-item"], [class*="list-item"], [class*="专业"], [class*="major"]'
            );

            items.forEach(item => {
                const text = item.textContent || '';
                if (text.length < 10) return;

                const scoreMatch = text.match(/(\\d{2,3})分/);
                const majorMatch = text.match(/([^\\s\\n，。；]{2,}(?:专业|学|工程|技术|管理|经济|法学|文学|理学|工学|医学|农学|军事|艺术|教育|历史|哲学)[^\\s\\n，。；]*)/);
                const rankMatch = text.match(/(\\d+)位次/);

                if (scoreMatch) {
                    const record = {
                        school_name: schoolName,
                        major_name: majorMatch ? majorMatch[1] : text.substring(0, 20),
                        min_score: parseInt(scoreMatch[1]),
                        province: province,
                        year: year,
                        batch: batch,
                        source: '夸克高考',
                    };

                    if (rankMatch) {
                        record.min_rank = parseInt(rankMatch[1]);
                    }

                    results.push(record);
                }
            });
        }

        // 方法3: 查找所有包含分数的元素
        if (results.length === 0) {
            const allElements = document.querySelectorAll('*');
            const seen = new Set();

            allElements.forEach(el => {
                const text = (el.textContent || '').trim();
                if (text.length < 5 || text.length > 200) return;
                if (seen.has(text)) return;

                const scoreMatch = text.match(/最低分[：:]?\\s*(\\d{2,3})/);
                const majorMatch = text.match(/^(.+?)\\s*\\d+分/);

                if (scoreMatch) {
                    seen.add(text);
                }
            });
        }

        return results;
    }

    return extractMajorScores(arguments[0], arguments[1], arguments[2], arguments[3]);
    '''

    try:
        results = driver.execute_script(script, school_name, year, PROVINCE, BATCH)
        return results if results else []
    except Exception as e:
        print(f'  JS提取出错: {e}')
        return []


def collect_network_responses(driver):
    """收集网络响应"""
    script = '''
    if (!window.__networkResponses) {
        window.__networkResponses = [];

        const origFetch = window.fetch;
        window.fetch = function(...args) {
            return origFetch.apply(this, args).then(response => {
                const url = args[0];
                if (typeof url === 'string' && 
                    (url.includes('score') || url.includes('major') || url.includes('fen_shu') || url.includes('luqu') || url.includes('college'))) {
                    response.clone().json().then(data => {
                        window.__networkResponses.push({ url, data });
                    }).catch(() => {});
                }
                return response;
            });
        };

        const origXHR = window.XMLHttpRequest;
        window.XMLHttpRequest = function() {
            const xhr = new origXHR();
            let _url = '';
            const origOpen = xhr.open;
            xhr.open = function(method, url) {
                _url = url;
                return origOpen.apply(this, arguments);
            };
            xhr.addEventListener('load', function() {
                if ((_url.includes('score') || _url.includes('major') || _url.includes('fen_shu')) &&
                    xhr.status === 200) {
                    try {
                        const data = JSON.parse(xhr.responseText);
                        window.__networkResponses.push({ url: _url, data });
                    } catch(e) {}
                }
            });
            return xhr;
        };
    }
    return window.__networkResponses ? window.__networkResponses.length : 0;
    '''

    try:
        count = driver.execute_script(script)
        return count
    except Exception:
        return 0


def extract_from_network_data(driver, school_name, year):
    """从网络数据中提取"""
    script = '''
    function extractFromNetwork(schoolName, year, province, batch) {
        const responses = window.__networkResponses || [];
        const results = [];

        function processItem(item) {
            const majorName = item.major_name || item.majorName || item.name || item.major || '';
            if (!majorName) return;

            results.push({
                school_name: schoolName,
                school_code: item.school_code || item.schoolCode || '',
                province: item.province_name || item.province || province,
                level: item.school_type || item.level || '',
                major_name: majorName,
                major_group: item.major_group_name || item.majorGroup || item.groupName || '',
                subject_requirement: item.subject_requirement || item.subjectRequirement || item.subject || '',
                year: year,
                min_score: item.min_score ?? item.minScore ?? item.score ?? null,
                min_rank: item.min_rank ?? item.minRank ?? item.rank ?? null,
                avg_score: item.avg_score ?? item.avgScore ?? null,
                batch: item.batch_name || item.batch || item.batchName || batch,
                person_count: item.person_count ?? item.personCount ?? item.plan_num ?? null,
                source: '夸克高考',
            });
        }

        function traverse(obj) {
            if (!obj || typeof obj !== 'object') return;
            if (Array.isArray(obj)) {
                obj.forEach(traverse);
                return;
            }

            if (obj.major_name || obj.majorName || (obj.name && (obj.min_score || obj.minScore))) {
                processItem(obj);
            }

            if (obj.list && Array.isArray(obj.list)) {
                obj.list.forEach(traverse);
            }
            if (obj.data) traverse(obj.data);
            if (obj.result) traverse(obj.result);

            Object.values(obj).forEach(val => {
                if (val && typeof val === 'object') traverse(val);
            });
        }

        responses.forEach(resp => traverse(resp.data));
        return results;
    }

    return extractFromNetwork(arguments[0], arguments[1], arguments[2], arguments[3]);
    '''

    try:
        results = driver.execute_script(script, school_name, year, PROVINCE, BATCH)
        return results if results else []
    except Exception as e:
        print(f'  网络数据提取出错: {e}')
        return []


def scroll_page(driver):
    """滚动页面以加载更多内容"""
    try:
        last_height = driver.execute_script('return document.body.scrollHeight')
        for _ in range(5):
            driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')
            time.sleep(2)
            new_height = driver.execute_script('return document.body.scrollHeight')
            if new_height == last_height:
                break
            last_height = new_height

        driver.execute_script('window.scrollTo(0, 0);')
        time.sleep(1)
    except Exception:
        pass


def main():
    print('=' * 70)
    print('夸克高考专业分数线自动抓取工具 v2')
    print('=' * 70)

    ensure_dir(OUTPUT_DIR)

    schools = read_schools_from_excel()
    if not schools:
        print('未找到院校数据，使用测试院校')
        schools = ['清华大学']

    driver = setup_driver()

    all_records = []
    success_count = 0
    fail_count = 0

    try:
        test_school = schools[0]
        print(f'\n测试院校: {test_school}')

        for year in YEARS:
            print(f'\n  --- {year}年 ---')

            url = build_school_url(test_school, year)
            print(f'  访问页面...')

            try:
                driver.get(url)
            except Exception as e:
                print(f'  页面加载超时: {e}')

            print('  等待页面加载...')
            wait_for_page_load(driver, timeout=15)

            print('  注入网络拦截...')
            collect_network_responses(driver)

            print('  刷新页面以捕获请求...')
            try:
                driver.refresh()
                wait_for_page_load(driver, timeout=15)
            except Exception:
                pass

            print('  滚动页面加载内容...')
            scroll_page(driver)
            time.sleep(3)

            print('  尝试从网络数据提取...')
            records = extract_from_network_data(driver, test_school, year)
            print(f'    网络数据提取到 {len(records)} 条')

            if not records:
                print('  尝试从DOM提取...')
                records = extract_data_via_js(driver, test_school, year)
                print(f'    DOM提取到 {len(records)} 条')

            if records:
                all_records.extend(records)
                success_count += 1

                year_file = os.path.join(OUTPUT_DIR, f'major_scores_{year}.json')
                with open(year_file, 'w', encoding='utf-8') as f:
                    json.dump(records, f, ensure_ascii=False, indent=2)
                print(f'  已保存到: {year_file}')
            else:
                fail_count += 1

            screenshot_file = os.path.join(OUTPUT_DIR, f'screenshot_{year}.png')
            try:
                driver.save_screenshot(screenshot_file)
                print(f'  截图已保存')
            except Exception:
                pass

            html_file = os.path.join(OUTPUT_DIR, f'page_{year}.html')
            try:
                with open(html_file, 'w', encoding='utf-8') as f:
                    f.write(driver.page_source)
                print(f'  页面HTML已保存')
            except Exception:
                pass

            time.sleep(2)

        all_file = os.path.join(OUTPUT_DIR, 'major_scores_all.json')
        with open(all_file, 'w', encoding='utf-8') as f:
            json.dump(all_records, f, ensure_ascii=False, indent=2)

        print('\n' + '=' * 70)
        print('测试抓取完成！')
        print(f'共获取 {len(all_records)} 条专业分数线数据')
        print(f'成功: {success_count} 年 | 失败: {fail_count} 年')
        print(f'数据已保存到: {all_file}')
        print('=' * 70)

        if all_records:
            print('\n数据预览:')
            for i, record in enumerate(all_records[:5]):
                print(f'  {i+1}. {record.get("school_name")} - {record.get("major_name")}')
                print(f'     分数: {record.get("min_score")} | 位次: {record.get("min_rank")} | 年份: {record.get("year")}')

        print('\n浏览器将在60秒后自动关闭...')
        time.sleep(60)

    except Exception as e:
        print(f'抓取出错: {e}')
        import traceback
        traceback.print_exc()
        time.sleep(30)
    finally:
        try:
            driver.quit()
        except Exception:
            pass
        print('浏览器已关闭')


if __name__ == '__main__':
    main()
