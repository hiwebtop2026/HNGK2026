from dotenv import load_dotenv
load_dotenv()

# -*- coding: utf-8 -*-
"""
夸克高考专业分数线自动抓取工具
使用Selenium + Edge浏览器自动化抓取数据
"""

import json
import time
import os
import re
from datetime import datetime

from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.microsoft import EdgeChromiumDriverManager

import openpyxl
import requests

EXCEL_FILE = r'C:\Users\lhp\Desktop\2023-2025年海南高考本科投档分数线.xlsx'
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data')

YEARS = [2025, 2024, 2023]
PROVINCE = '海南'
BATCH = '本科批'

SUPABASE_URL = os.environ.get('SUPABASE_URL')
SUPABASE_ANON_KEY = os.environ.get('SUPABASE_ANON_KEY')


def ensure_dir(dir_path):
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)


def read_schools_from_excel():
    """从Excel读取院校列表"""
    print('正在读取Excel文件获取院校列表...')
    schools = set()

    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)

        for year in ['2023', '2024', '2025']:
            if year not in wb.sheetnames:
                continue

            ws = wb[year]

            name_col_idx = None
            row_iter = ws.iter_rows(min_row=2, max_row=2, values_only=True)
            for row in row_iter:
                for idx, cell in enumerate(row):
                    if cell and isinstance(cell, str) and ('名称' in cell or '院校' in cell):
                        name_col_idx = idx
                        break

            if name_col_idx is None:
                continue

            for row in ws.iter_rows(min_row=3, values_only=True):
                if len(row) <= name_col_idx:
                    continue
                cell_val = row[name_col_idx]
                if cell_val and isinstance(cell_val, str) and len(cell_val) >= 2:
                    school_name = re.sub(r'\([^)]*\)', '', cell_val).strip()
                    school_name = re.sub(r'\d+$', '', school_name).strip()
                    if len(school_name) >= 2:
                        schools.add(school_name)

        wb.close()
    except Exception as e:
        print(f'读取Excel失败: {e}')

    result = sorted(list(schools))
    print(f'共提取 {len(result)} 个院校')
    return result


def setup_driver():
    """设置Edge浏览器"""
    print('正在启动Edge浏览器...')

    options = Options()
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_argument('--start-maximized')
    options.add_argument('--disable-infobars')
    options.add_argument('--disable-gpu')
    options.add_experimental_option('excludeSwitches', ['enable-automation'])
    options.add_experimental_option('useAutomationExtension', False)

    try:
        driver = webdriver.Edge(options=options)
    except Exception:
        print('正在下载Edge驱动...')
        service = Service(EdgeChromiumDriverManager().install())
        driver = webdriver.Edge(service=service, options=options)

    driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
        'source': '''
            Object.defineProperty(navigator, 'webdriver', {
                get: () => undefined
            });
            window.chrome = { runtime: {} };
        '''
    })

    return driver


def build_school_url(school_name, year):
    """构建院校页面URL"""
    params = json.dumps({
        'province': PROVINCE,
        'year': str(year),
        'batch': BATCH,
        'genre': '综合'
    }, ensure_ascii=False)

    base_url = 'https://vt.quark.cn/blm/gaokao-college-794/tab'
    url = (f'{base_url}?app=fen_shu_xian'
           f'&university_name={school_name}'
           f'&q={school_name}'
           f'&device=pc'
           f'&by=tuijian'
           f'&by2=general_entity_college'
           f'&params={params}'
           f'&type=luqu')

    return url


def extract_data_from_dom(driver, school_name, year):
    """从DOM提取专业分数线数据"""
    results = []

    try:
        tables = driver.find_elements(By.TAG_NAME, 'table')
        for table in tables:
            rows = table.find_elements(By.TAG_NAME, 'tr')
            if len(rows) < 2:
                continue

            headers = []
            first_row = rows[0].find_elements(By.TAG_NAME, 'th')
            if not first_row:
                first_row = rows[0].find_elements(By.TAG_NAME, 'td')
            for cell in first_row:
                headers.append(cell.text.strip())

            for row in rows[1:]:
                cells = row.find_elements(By.TAG_NAME, 'td')
                if len(cells) < 2:
                    continue

                record = {
                    'school_name': school_name,
                    'province': PROVINCE,
                    'year': year,
                    'batch': BATCH,
                    'source': '夸克高考',
                }

                for idx, cell in enumerate(cells):
                    if idx >= len(headers):
                        break
                    header = headers[idx]
                    value = cell.text.strip()

                    if '专业' in header and '组' not in header:
                        record['major_name'] = value
                    elif '专业组' in header or '组名' in header:
                        record['major_group'] = value
                    elif '最低分' in header or ('分' in header and '位次' not in header and '排名' not in header and '差' not in header):
                        try:
                            num = int(re.sub(r'[^\d]', '', value))
                            record['min_score'] = num
                        except (ValueError, TypeError):
                            pass
                    elif '位次' in header or '排名' in header:
                        try:
                            num = int(re.sub(r'[^\d]', '', value))
                            record['min_rank'] = num
                        except (ValueError, TypeError):
                            pass
                    elif '平均分' in header or '均分' in header:
                        try:
                            num = int(re.sub(r'[^\d]', '', value))
                            record['avg_score'] = num
                        except (ValueError, TypeError):
                            pass
                    elif '科目' in header or '选科' in header or '要求' in header:
                        record['subject_requirement'] = value
                    elif '人数' in header or '计划' in header:
                        try:
                            num = int(re.sub(r'[^\d]', '', value))
                            record['person_count'] = num
                        except (ValueError, TypeError):
                            pass

                if record.get('major_name') and (record.get('min_score') or record.get('min_rank')):
                    results.append(record)

    except Exception as e:
        print(f'  DOM提取出错: {e}')

    if not results:
        try:
            script = '''
            const results = [];
            const items = document.querySelectorAll('[class*="major"], [class*="专业"], [class*="score-item"], [class*="list-item"]');
            items.forEach(item => {
                const text = item.textContent || '';
                if (text.length > 10 && text.match(/\\d{2,3}分/)) {
                    const majorMatch = text.match(/([^\\s\\d]{2,}(?:专业|学|工程|技术|管理|经济|法学|文学|理学|工学|医学|农学|军事|艺术|教育|历史|哲学)[^\\s]*)/);
                    const scoreMatch = text.match(/(\\d{2,3})分/);
                    const rankMatch = text.match(/(\\d+)位次/);
                    if (scoreMatch) {
                        results.push({
                            major_name: majorMatch ? majorMatch[1] : '',
                            min_score: parseInt(scoreMatch[1]),
                            min_rank: rankMatch ? parseInt(rankMatch[1]) : null
                        });
                    }
                }
            });
            return results;
            '''
            js_results = driver.execute_script(script)
            if js_results:
                for item in js_results:
                    item['school_name'] = school_name
                    item['province'] = PROVINCE
                    item['year'] = year
                    item['batch'] = BATCH
                    item['source'] = '夸克高考'
                    results.append(item)
        except Exception as e:
            print(f'  JS提取出错: {e}')

    return results


def click_show_all(driver):
    """点击查看全部按钮"""
    texts_to_try = ['查看全部', '全部', '更多', '展开全部', '显示全部', '查看更多', '加载更多']

    for text in texts_to_try:
        try:
            elements = driver.find_elements(By.XPATH, f'//*[contains(text(), "{text}")]')
            for el in elements:
                if el.is_displayed() and el.is_enabled():
                    try:
                        el.click()
                        time.sleep(2)
                        print(f'  点击了: {text}')
                        return True
                    except Exception:
                        try:
                            driver.execute_script('arguments[0].click();', el)
                            time.sleep(2)
                            print(f'  点击了(JS): {text}')
                            return True
                        except Exception:
                            continue
        except Exception:
            continue

    return False


def switch_year(driver, year):
    """切换年份"""
    year_str = str(year)

    try:
        elements = driver.find_elements(By.XPATH, f'//*[text()="{year_str}"]')
        for el in elements:
            try:
                if el.is_displayed():
                    el.click()
                    time.sleep(2)
                    return True
            except Exception:
                try:
                    driver.execute_script('arguments[0].click();', el)
                    time.sleep(2)
                    return True
                except Exception:
                    continue
    except Exception:
        pass

    return False


def save_to_supabase(records):
    """保存到Supabase"""
    if not records:
        return 0

    try:
        batch_size = 100
        inserted = 0

        for i in range(0, len(records), batch_size):
            batch = records[i:i + batch_size]
            url = f'{SUPABASE_URL}/rest/v1/major_scores'
            headers = {
                'apikey': SUPABASE_ANON_KEY,
                'Authorization': f'Bearer {SUPABASE_ANON_KEY}',
                'Content-Type': 'application/json',
                'Prefer': 'return=minimal'
            }

            response = requests.post(url, json=batch, headers=headers)
            if response.status_code in [200, 201]:
                inserted += len(batch)
            else:
                print(f'  Supabase插入失败: {response.status_code} {response.text[:200]}')
                break

        return inserted
    except Exception as e:
        print(f'  保存到Supabase失败: {e}')
        return 0


def main():
    print('=' * 70)
    print('夸克高考专业分数线自动抓取工具')
    print('=' * 70)

    ensure_dir(OUTPUT_DIR)

    schools = read_schools_from_excel()
    if not schools:
        print('未找到院校数据，退出')
        return

    driver = setup_driver()

    all_records = []
    success_count = 0
    fail_count = 0

    try:
        test_school = schools[0]
        print(f'\n测试院校: {test_school}')

        for year in YEARS:
            print(f'\n  --- {year}年 ---')

            url = build_school_url(test_school, year)
            print(f'  访问: {url[:80]}...')

            try:
                driver.get(url)
                time.sleep(5)
            except Exception as e:
                print(f'  页面加载失败: {e}')
                fail_count += 1
                continue

            click_show_all(driver)
            time.sleep(2)

            records = extract_data_from_dom(driver, test_school, year)
            print(f'  提取到 {len(records)} 条数据')

            if records:
                all_records.extend(records)
                success_count += 1

                year_file = os.path.join(OUTPUT_DIR, f'major_scores_{year}.json')
                with open(year_file, 'w', encoding='utf-8') as f:
                    json.dump(records, f, ensure_ascii=False, indent=2)
                print(f'  已保存到: {year_file}')
            else:
                fail_count += 1

            screenshot_file = os.path.join(OUTPUT_DIR, f'screenshot_{year}.png')
            try:
                driver.save_screenshot(screenshot_file)
                print(f'  截图已保存: {screenshot_file}')
            except Exception:
                pass

            time.sleep(1)

        all_file = os.path.join(OUTPUT_DIR, 'major_scores_all.json')
        with open(all_file, 'w', encoding='utf-8') as f:
            json.dump(all_records, f, ensure_ascii=False, indent=2)

        print('\n' + '=' * 70)
        print('测试抓取完成！')
        print(f'共获取 {len(all_records)} 条专业分数线数据')
        print(f'成功: {success_count} 年 | 失败: {fail_count} 年')
        print(f'数据已保存到: {all_file}')
        print('=' * 70)

        if all_records:
            print('\n数据预览:')
            for i, record in enumerate(all_records[:5]):
                print(f'  {i+1}. {record.get("school_name")} - {record.get("major_name")}')
                print(f'     分数: {record.get("min_score")} | 位次: {record.get("min_rank")} | 年份: {record.get("year")}')

        input('\n按Enter键关闭浏览器...')

    except Exception as e:
        print(f'抓取出错: {e}')
        import traceback
        traceback.print_exc()
    finally:
        driver.quit()
        print('浏览器已关闭')


if __name__ == '__main__':
    main()
