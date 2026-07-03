import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

def determine_tier(row, student_score, student_rank):
    min_score = row.get('min_score')
    min_rank = row.get('min_rank')
    
    if pd.isna(min_score) and pd.isna(min_rank):
        return None
    
    score_diff = 0
    rank_diff = 0
    score_valid = False
    rank_valid = False
    
    if not pd.isna(min_score):
        try:
            score_diff = student_score - float(min_score)
            score_valid = True
        except:
            score_diff = 0
    
    if not pd.isna(min_rank):
        try:
            rank_diff = float(min_rank) - student_rank
            rank_valid = True
        except:
            rank_diff = 0
    
    if not score_valid and not rank_valid:
        return None
    
    tier_score = None
    tier_rank = None
    
    if score_valid:
        if score_diff >= 15:
            tier_score = '保'
        elif score_diff >= 5:
            tier_score = '稳'
        elif score_diff >= -10:
            tier_score = '冲'
    
    if rank_valid:
        if rank_diff >= 3000:
            tier_rank = '保'
        elif rank_diff >= 1000:
            tier_rank = '稳'
        elif rank_diff >= -2000:
            tier_rank = '冲'
    
    if tier_score and tier_rank:
        if tier_score == tier_rank:
            return tier_score
        else:
            if tier_score == '保' or tier_rank == '保':
                return '保'
            elif tier_score == '稳' or tier_rank == '稳':
                return '稳'
            else:
                return '冲'
    elif tier_score:
        return tier_score
    elif tier_rank:
        return tier_rank
    else:
        return None

def main():
    input_file = r'G:\Downloads\20260703.xlsx'
    output_file = r'G:\Downloads\20260703_志愿匹配.xlsx'
    
    student_score = 603
    student_rank = 12085
    
    print(f"处理文件: {input_file}")
    print(f"考生信息: 分数={student_score}, 位次={student_rank}")
    
    try:
        df = pd.read_excel(input_file)
        print(f"成功读取文件，共 {len(df)} 行数据")
        print(f"列名: {list(df.columns)}")
        
        if 'min_score' not in df.columns and 'min_rank' not in df.columns:
            print("错误：文件中没有找到 min_score 或 min_rank 列")
            return
        
        df['志愿匹配'] = df.apply(lambda row: determine_tier(row, student_score, student_rank), axis=1)
        
        df.to_excel(output_file, index=False)
        print(f"数据处理完成，已保存到: {output_file}")
        
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        
        tier_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == '志愿匹配':
                tier_col = col
                break
        
        if tier_col:
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=tier_col)
                value = cell.value
                if value == '保':
                    cell.fill = green_fill
                    cell.font = Font(color='FFFFFF', bold=True)
                elif value == '稳':
                    cell.fill = yellow_fill
                    cell.font = Font(color='000000', bold=True)
                elif value == '冲':
                    cell.fill = red_fill
                    cell.font = Font(color='FFFFFF', bold=True)
            
            col_letter = get_column_letter(tier_col)
            ws.column_dimensions[col_letter].width = 12
            
            wb.save(output_file)
            print(f"已为志愿匹配列添加颜色标记")
        else:
            print("未找到志愿匹配列")
        
        chong_count = df['志愿匹配'].value_counts().get('冲', 0)
        wen_count = df['志愿匹配'].value_counts().get('稳', 0)
        bao_count = df['志愿匹配'].value_counts().get('保', 0)
        
        print(f"\n匹配结果统计:")
        print(f"  冲刺志愿: {chong_count} 个")
        print(f"  稳妥志愿: {wen_count} 个")
        print(f"  保底志愿: {bao_count} 个")
        print(f"  总计: {chong_count + wen_count + bao_count} 个")
        
    except Exception as e:
        print(f"处理过程中发生错误: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()