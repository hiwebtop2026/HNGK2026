# -*- coding: utf-8 -*-
"""
夸克高考专业分数线自动抓取工具 (CDP连接本地Edge)
"""

import json
import time
import os
import re
import sys

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
except ImportError:
    os.system('pip install selenium')
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC

try:
    import openpyxl
except ImportError:
    os.system('pip install openpyxl')
    import openpyxl

EXCEL_FILE = r'C:\Users\lhp\Desktop\2023-2025年海南高考本科投档分数线.xlsx'
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data')

YEARS = [2025, 2024, 2023]
PROVINCE = '海南'
BATCH = '本科批'
GENRE = '综合'

EDGE_DEBUG_PORT = 9222

def ensure_dir(dir_path):
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)

def read_schools_from_excel():
    print('正在读取Excel文件获取院校列表...')
    schools = set()

    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)

        for year in ['2025', '2024', '2023']:
            if year not in wb.sheetnames:
                continue

            ws = wb[year]

            name_col_idx = None
            header_row_idx = 2

            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
                for col_idx, cell in enumerate(row):
                    if cell and isinstance(cell, str):
                        cell_str = cell.strip()
                        if ('院校' in cell_str and '名称' in cell_str) or cell_str == '院校专业组名称':
                            name_col_idx = col_idx
                            header_row_idx = row_idx
                            break
                if name_col_idx is not None:
                    break

            if name_col_idx is None:
                continue

            count = 0
            for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                if len(row) <= name_col_idx:
                    continue

                cell_val = row[name_col_idx]
                if not cell_val or not isinstance(cell_val, str):
                    continue

                cell_val = cell_val.strip()
                if len(cell_val) < 2:
                    continue

                if '科目' in cell_val and '要求' in cell_val:
                    continue
                if cell_val in ['说明', '注', '备注']:
                    continue

                school_name = re.sub(r'\([^)]*\)', '', cell_val).strip()
                school_name = re.sub(r'\d+$', '', school_name).strip()

                if len(school_name) >= 2 and not any(c.isdigit() for c in school_name[:2]):
                    schools.add(school_name)
                    count += 1

            print(f'  {year}年: {count}个')

        wb.close()
    except Exception as e:
        print(f'读取Excel失败: {e}')

    result = sorted(list(schools))
    print(f'共提取 {len(result)} 个唯一院校')
    return result

def connect_to_edge_cdp():
    print(f'正在连接到Edge浏览器 CDP端口: {EDGE_DEBUG_PORT}...')
    
    try:
        options = webdriver.EdgeOptions()
        options.add_experimental_option('debuggerAddress', f'localhost:{EDGE_DEBUG_PORT}')
        driver = webdriver.Edge(options=options)
        print('✅ 连接成功！')
        return driver
    except Exception as e:
        print(f'❌ 连接失败: {e}')
        print('')
        print('请按以下步骤操作:')
        print('')
        print('1. 关闭所有Edge浏览器窗口')
        print('2. 按 Win+R 打开运行窗口')
        print('3. 输入以下命令并回车:')
        print(f'   msedge --remote-debugging-port={EDGE_DEBUG_PORT}')
        print('')
        print('4. 在打开的Edge浏览器中访问夸克高考页面')
        print('   例如: https://vt.quark.cn/blm/gaokao-college-794/tab')
        print('')
        print('5. 重新运行此脚本')
        return None

def build_school_url(school_name, year):
    params = json.dumps({
        'province': PROVINCE,
        'year': str(year),
        'batch': BATCH,
        'genre': GENRE,
    }, ensure_ascii=False)

    return (
        'https://vt.quark.cn/blm/gaokao-college-794/tab?app=fen_shu_xian'
        f'&university_name={school_name}'
        f'&q={school_name}'
        '&uc_biz_str=qk_enable_gesture%3Atrue%7COPT%3AW_ENTER_ANI%401%7COPT%3ATOOLBAR_STYLE%400%7COPT%3AW_PAGE_REFRESH%400%7COPT%3ABACK_BTN_STYLE%400%7COPT%3AIMMERSIVE%401%7COPT%3AW_PAGE_REFRESH%400'
        '&device=pc'
        '&bar=pure'
        '&by=tuijian'
        '&by2=general_entity_college'
        f'&params={params}'
        '&type=luqu'
        '&from='
        '&uc_param_str=ntnwvepffrbiprsvchutosstxskp'
    )

def switch_to_major_tab(driver):
    print('  切换到专业分数线Tab...')
    
    script = '''
    const tabs = document.querySelectorAll('[class*="tab"], [role="tab"]');
    for (const tab of tabs) {
        const text = (tab.textContent || '').trim();
        if (text.includes('专业分数线') || text === '专业分数线') {
            tab.click();
            return { clicked: true, text: text };
        }
    }
    
    const allClickable = document.querySelectorAll('div, span, button, a');
    for (const el of allClickable) {
        const text = (el.textContent || '').trim();
        if (text === '专业分数线' || (text.length <= 10 && text.includes('专业') && text.includes('分数线'))) {
            const style = window.getComputedStyle(el);
            if (style.cursor === 'pointer' || el.onclick || el.closest('[role="tab"]')) {
                el.click();
                return { clicked: true, text: text };
            }
        }
    }
    
    return { clicked: false };
    '''
    
    try:
        result = driver.execute_script(script)
        if result and result.get('clicked'):
            print(f'    已切换到: {result.get("text")}')
            time.sleep(3)
            return True
    except Exception as e:
        print(f'  切换Tab失败: {e}')
    
    print('    未找到专业分数线Tab')
    return False

def click_view_all(driver):
    print('  点击查看全部...')
    
    script = '''
    const allBtns = document.querySelectorAll('button, a, div, span');
    for (const btn of allBtns) {
        const text = (btn.textContent || '').trim();
        if (text === '查看全部' || text === '全部') {
            const style = window.getComputedStyle(btn);
            if (style.cursor === 'pointer' || btn.onclick || btn.tagName === 'BUTTON' || btn.tagName === 'A') {
                btn.click();
                return { clicked: true, text: text };
            }
        }
    }
    
    const moreBtns = document.querySelectorAll('[class*="more"], [class*="all"], [class*="More"], [class*="All"]');
    for (const btn of moreBtns) {
        const text = (btn.textContent || '').trim();
        if (text.includes('全部') || text.includes('more')) {
            btn.click();
            return { clicked: true, text: text || 'more button' };
        }
    }
    
    return { clicked: false };
    '''
    
    try:
        result = driver.execute_script(script)
        if result and result.get('clicked'):
            print(f'    点击了: {result.get("text")}')
            time.sleep(3)
            return True
    except Exception as e:
        print(f'  点击查看全部失败: {e}')
    
    print('    未找到查看全部按钮')
    return False

def extract_major_scores(driver, school_name, year):
    print('  提取专业分数线数据...')
    
    network_data = extract_from_network(driver, school_name, year)
    if network_data:
        print(f'    从网络请求获取到 {len(network_data)} 条数据')
        return network_data
    
    dom_data = extract_from_dom(driver, school_name, year)
    if dom_data:
        print(f'    从DOM提取到 {len(dom_data)} 条数据')
        return dom_data
    
    print('    未提取到数据')
    return []

def extract_from_network(driver, school_name, year):
    return []

def extract_from_dom(driver, school_name, year):
    script = '''
    function extractData(schoolName, year, province, batch) {
        const results = [];
        
        const allLists = document.querySelectorAll('[class*="list"], [class*="List"], ul, ol');
        
        for (const list of allLists) {
            const items = list.querySelectorAll('li, [class*="item"], div');
            
            for (const item of items) {
                const text = item.textContent || '';
                if (text.length < 5 || text.length > 500) continue;
                
                const majorMatch = text.match(/([^\s\d\n]{2,}(?:专业|学|工程|技术|管理|经济|法学|文学|理学|工学|医学|农学|军事|艺术|教育|历史|哲学|科学)[^\s\n]*)/);
                const scoreMatch = text.match(/(\d{2,3})\s*分/);
                const rankMatch = text.match(/(\d{3,6})\s*位次/);
                const countMatch = text.match(/(\d+)\s*人/);
                
                if (scoreMatch && majorMatch) {
                    const record = {
                        school_name: schoolName,
                        major_name: majorMatch[1],
                        min_score: parseInt(scoreMatch[1]),
                        province: province,
                        year: year,
                        batch: batch,
                        source: '夸克高考',
                    };
                    
                    if (rankMatch) {
                        record.min_rank = parseInt(rankMatch[1]);
                    }
                    if (countMatch) {
                        record.person_count = parseInt(countMatch[1]);
                    }
                    
                    results.push(record);
                }
            }
        }
        
        const tables = document.querySelectorAll('table');
        for (const table of tables) {
            const rows = table.querySelectorAll('tr');
            if (rows.length < 2) continue;
            
            const headers = [];
            rows[0].querySelectorAll('th, td').forEach(cell => {
                headers.push((cell.textContent || '').trim());
            });
            
            for (let i = 1; i < rows.length; i++) {
                const cells = rows[i].querySelectorAll('td');
                if (cells.length < 2) continue;
                
                const record = {
                    school_name: schoolName,
                    province: province,
                    year: year,
                    batch: batch,
                    source: '夸克高考',
                };
                
                cells.forEach((cell, idx) => {
                    if (idx >= headers.length) return;
                    const header = headers[idx];
                    const value = (cell.textContent || '').trim();
                    
                    if (header.includes('专业') && !header.includes('组')) {
                        record.major_name = value;
                    } else if (header.includes('专业组')) {
                        record.major_group = value;
                    } else if (header.includes('最低分') || (header.includes('分') && !header.includes('位次'))) {
                        const num = parseInt(value);
                        if (!isNaN(num)) record.min_score = num;
                    } else if (header.includes('位次') || header.includes('排名')) {
                        const num = parseInt(value);
                        if (!isNaN(num)) record.min_rank = num;
                    } else if (header.includes('科目') || header.includes('选科')) {
                        record.subject_requirement = value;
                    }
                });
                
                if (record.major_name && record.min_score) {
                    results.push(record);
                }
            }
        }
        
        const uniqueResults = [];
        const seen = new Set();
        for (const r of results) {
            const key = r.major_name + '_' + r.min_score;
            if (!seen.has(key)) {
                seen.add(key);
                uniqueResults.push(r);
            }
        }
        
        return uniqueResults;
    }
    
    return extractData(arguments[0], arguments[1], arguments[2], arguments[3]);
    '''
    
    try:
        results = driver.execute_script(script, school_name, year, PROVINCE, BATCH)
        return results if results else []
    except Exception as e:
        print(f'  DOM提取出错: {e}')
        return []

def scrape_school(driver, school_name):
    print(f'\n处理院校: {school_name}')
    all_records = []

    for year in YEARS:
        print(f'\n  === {year}年 ===')
        
        url = build_school_url(school_name, year)
        print(f'  访问: {url[:80]}...')
        
        try:
            driver.get(url)
            time.sleep(5)
        except Exception as e:
            print(f'  页面加载失败: {e}')
            continue

        time.sleep(3)

        switch_to_major_tab(driver)
        click_view_all(driver)
        time.sleep(2)

        records = extract_major_scores(driver, school_name, year)
        print(f'  提取到 {len(records)} 条数据')

        if records:
            all_records.extend(records)
            
            print('  数据预览:')
            for i, r in enumerate(records[:3]):
                print(f'    {i+1}. {r.get("major_name", "")} - {r.get("min_score", "--")}分 - {r.get("min_rank", "--")}位次')

        time.sleep(2)

    return all_records

def main():
    print('=' * 70)
    print('夸克高考专业分数线自动抓取工具 (CDP连接Edge)')
    print('=' * 70)

    ensure_dir(OUTPUT_DIR)

    schools = read_schools_from_excel()
    if not schools:
        print('未找到院校数据，退出')
        return

    driver = connect_to_edge_cdp()
    if not driver:
        return

    all_records = []
    success_count = 0
    fail_count = 0

    try:
        test_school = '华东理工大学'
        print(f'\n测试院校: {test_school}')

        records = scrape_school(driver, test_school)
        
        if records:
            all_records.extend(records)
            success_count += 1
            
            test_file = os.path.join(OUTPUT_DIR, 'test_major_scores.json')
            with open(test_file, 'w', encoding='utf-8') as f:
                json.dump(records, f, ensure_ascii=False, indent=2)
            print(f'\n测试数据已保存到: {test_file}')
            
            print('\n✅ 测试成功！')
            print('如需批量抓取所有院校，请修改脚本中的 BATCH_SCRAPE = True')
        else:
            fail_count += 1
            print('\n❌ 测试抓取失败')

        all_file = os.path.join(OUTPUT_DIR, 'major_scores_all.json')
        with open(all_file, 'w', encoding='utf-8') as f:
            json.dump(all_records, f, ensure_ascii=False, indent=2)

        print('\n' + '=' * 70)
        print('测试抓取完成！')
        print(f'共获取 {len(all_records)} 条专业分数线数据')
        print(f'成功: {success_count} | 失败: {fail_count}')
        print(f'数据已保存到: {all_file}')
        print('=' * 70)

        print('\n浏览器保持打开，按任意键退出...')
        try:
            input()
        except:
            time.sleep(30)

    except Exception as e:
        print(f'抓取出错: {e}')
        import traceback
        traceback.print_exc()
        time.sleep(30)
    finally:
        try:
            driver.quit()
        except Exception:
            pass
        print('浏览器已关闭')

if __name__ == '__main__':
    main()
