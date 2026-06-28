# -*- coding: utf-8 -*-
"""
生成包含院校列表的完整浏览器抓取脚本
"""

import os
import re
import json

try:
    import openpyxl
except ImportError:
    os.system('pip install openpyxl')
    import openpyxl

EXCEL_FILE = r'C:\Users\lhp\Desktop\2023-2025年海南高考本科投档分数线.xlsx'
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data')
SCRAPER_JS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'browser_auto_scraper.js')
OUTPUT_JS_FILE = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data', 'quark_scraper_complete.js')

def read_schools_from_excel():
    print('正在读取Excel文件获取院校列表...')
    schools = set()

    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)

        for year in ['2025', '2024', '2023']:
            if year not in wb.sheetnames:
                continue

            ws = wb[year]

            name_col_idx = None
            header_row_idx = 2

            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
                for col_idx, cell in enumerate(row):
                    if cell and isinstance(cell, str):
                        cell_str = cell.strip()
                        if ('院校' in cell_str and '名称' in cell_str) or cell_str == '院校专业组名称':
                            name_col_idx = col_idx
                            header_row_idx = row_idx
                            break
                if name_col_idx is not None:
                    break

            if name_col_idx is None:
                continue

            count = 0
            for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                if len(row) <= name_col_idx:
                    continue

                cell_val = row[name_col_idx]
                if not cell_val or not isinstance(cell_val, str):
                    continue

                cell_val = cell_val.strip()
                if len(cell_val) < 2:
                    continue

                if '科目' in cell_val and '要求' in cell_val:
                    continue
                if cell_val in ['说明', '注', '备注']:
                    continue

                school_name = re.sub(r'\([^)]*\)', '', cell_val).strip()
                school_name = re.sub(r'\d+$', '', school_name).strip()

                if len(school_name) >= 2 and not any(c.isdigit() for c in school_name[:2]):
                    schools.add(school_name)
                    count += 1

            print(f'  {year}年: {count}个')

        wb.close()
    except Exception as e:
        print(f'读取Excel失败: {e}')

    result = sorted(list(schools))
    print(f'共提取 {len(result)} 个唯一院校')
    return result

def generate_complete_script(schools):
    print('\n正在生成完整脚本...')
    
    with open(SCRAPER_JS_FILE, 'r', encoding='utf-8') as f:
        scraper_js = f.read()
    
    schools_json = json.dumps(schools, ensure_ascii=False, indent=4)
    
    # 替换空的SCHOOLS数组
    schools_array_line = '  const SCHOOLS = [\n    // 这里会在运行时动态获取，或者用户手动添加\n  ];'
    new_schools_array = f'  const SCHOOLS = {schools_json};'
    
    complete_js = scraper_js.replace(schools_array_line, new_schools_array)
    
    # 在末尾添加自动运行的代码
    auto_run_code = '''

// ============================================================
// 一键运行：直接运行 QuarkScraper.test() 测试，没问题再运行 QuarkScraper.start()
// ============================================================
console.log('%c🎉 院校列表已内置！共 ' + QuarkScraper.schools.length + ' 所院校', 'color: #22c55e; font-size: 14px; font-weight: bold;');
console.log('%c💡 快速开始：', 'color: #3b82f6; font-weight: bold;');
console.log('   1. 先运行 QuarkScraper.test() 测试当前页面');
console.log('   2. 测试没问题后，运行 QuarkScraper.start() 开始批量抓取');
console.log('   3. 抓取完成后会自动下载JSON文件');
'''
    
    complete_js = complete_js + auto_run_code
    
    with open(OUTPUT_JS_FILE, 'w', encoding='utf-8') as f:
        f.write(complete_js)
    
    print(f'✅ 完整脚本已生成: {OUTPUT_JS_FILE}')
    print(f'   文件大小: {os.path.getsize(OUTPUT_JS_FILE) / 1024:.1f} KB')

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    schools = read_schools_from_excel()
    if not schools:
        print('未找到院校数据，退出')
        return
    
    generate_complete_script(schools)
    
    print('\n' + '=' * 60)
    print('📋 使用说明:')
    print('=' * 60)
    print('')
    print('1. 在Edge浏览器中打开任意一个夸克高考院校的专业分数线页面')
    print('2. 按 F12 打开开发者工具')
    print('3. 切换到 Console (控制台) 标签')
    print(f'4. 打开文件: {OUTPUT_JS_FILE}')
    print('5. 把文件内容全部复制，粘贴到控制台，按回车运行')
    print('6. 运行 QuarkScraper.test() 测试一下')
    print('7. 测试没问题后，运行 QuarkScraper.start() 开始批量抓取')
    print('')
    print('=' * 60)

if __name__ == '__main__':
    main()
