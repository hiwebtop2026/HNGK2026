import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

INPUT_FILE = r'G:\Downloads\20260704.xlsx'
OUTPUT_FILE = r'G:\Downloads\海南高考志愿填报参考_优化版.xlsx'

SCHOOL_985 = [
    '北京大学', '清华大学', '复旦大学', '上海交通大学', '浙江大学',
    '南京大学', '中国科学技术大学', '哈尔滨工业大学', '西安交通大学',
    '北京航空航天大学', '北京理工大学', '中国人民大学', '北京师范大学',
    '武汉大学', '华中科技大学', '同济大学', '中山大学', '华南理工大学',
    '南开大学', '天津大学', '东南大学', '厦门大学', '山东大学',
    '四川大学', '吉林大学', '中南大学', '重庆大学', '西北工业大学',
    '兰州大学', '东北大学', '湖南大学', '大连理工大学', '华东师范大学',
    '电子科技大学', '中国农业大学', '中国海洋大学', '西北农林科技大学',
    '中央民族大学'
]

SCHOOL_211 = [
    '北京大学', '清华大学', '复旦大学', '上海交通大学', '浙江大学',
    '南京大学', '中国科学技术大学', '哈尔滨工业大学', '西安交通大学',
    '北京航空航天大学', '北京理工大学', '中国人民大学', '北京师范大学',
    '武汉大学', '华中科技大学', '同济大学', '中山大学', '华南理工大学',
    '南开大学', '天津大学', '东南大学', '厦门大学', '山东大学',
    '四川大学', '吉林大学', '中南大学', '重庆大学', '西北工业大学',
    '兰州大学', '东北大学', '湖南大学', '大连理工大学', '华东师范大学',
    '电子科技大学', '中国农业大学', '中国海洋大学', '西北农林科技大学',
    '中央民族大学', '北京交通大学', '北京工业大学', '北京科技大学',
    '北京化工大学', '北京邮电大学', '北京林业大学', '北京中医药大学',
    '北京外国语大学', '中国传媒大学', '中央财经大学', '对外经济贸易大学',
    '中国政法大学', '华北电力大学', '天津医科大学', '河北工业大学',
    '太原理工大学', '内蒙古大学', '辽宁大学', '大连海事大学',
    '延边大学', '东北师范大学', '哈尔滨工程大学', '东北农业大学',
    '东北林业大学', '华东理工大学', '东华大学', '上海外国语大学',
    '上海财经大学', '上海大学', '南京航空航天大学', '南京理工大学',
    '中国矿业大学', '南京邮电大学', '河海大学', '江南大学',
    '南京林业大学', '南京农业大学', '中国药科大学', '南京师范大学',
    '宁波大学', '安徽大学', '合肥工业大学', '福州大学', '南昌大学',
    '中国石油大学', '郑州大学', '中国地质大学', '武汉理工大学',
    '华中农业大学', '华中师范大学', '中南财经政法大学', '湖南师范大学',
    '暨南大学', '华南师范大学', '广西大学', '海南大学',
    '西南大学', '西南交通大学', '四川农业大学', '西南财经大学',
    '贵州大学', '云南大学', '西藏大学', '西北大学',
    '西安电子科技大学', '长安大学', '陕西师范大学', '青海大学',
    '宁夏大学', '新疆大学', '石河子大学'
]

SCHOOL_DOUBLE_FIRST = [
    '北京大学', '清华大学', '中国人民大学', '北京师范大学',
    '北京航空航天大学', '北京理工大学', '中国农业大学', '北京科技大学',
    '北京交通大学', '北京邮电大学', '北京化工大学', '北京林业大学',
    '北京中医药大学', '北京外国语大学', '中国传媒大学', '中央财经大学',
    '对外经济贸易大学', '中国政法大学', '华北电力大学', '南开大学',
    '天津大学', '天津医科大学', '河北工业大学', '太原理工大学',
    '内蒙古大学', '辽宁大学', '大连理工大学', '东北大学',
    '吉林大学', '延边大学', '东北师范大学', '哈尔滨工业大学',
    '哈尔滨工程大学', '东北农业大学', '东北林业大学', '复旦大学',
    '同济大学', '上海交通大学', '华东理工大学', '东华大学',
    '华东师范大学', '上海外国语大学', '上海财经大学', '上海大学',
    '南京大学', '东南大学', '南京航空航天大学', '南京理工大学',
    '中国矿业大学', '南京邮电大学', '河海大学', '江南大学',
    '南京农业大学', '中国药科大学', '南京师范大学', '浙江大学',
    '宁波大学', '安徽大学', '中国科学技术大学', '合肥工业大学',
    '厦门大学', '福州大学', '南昌大学', '山东大学',
    '中国海洋大学', '中国石油大学', '郑州大学', '武汉大学',
    '华中科技大学', '中国地质大学', '武汉理工大学', '华中农业大学',
    '华中师范大学', '中南财经政法大学', '湖南大学', '中南大学',
    '湖南师范大学', '中山大学', '暨南大学', '华南理工大学',
    '华南师范大学', '广西大学', '海南大学', '重庆大学',
    '西南大学', '四川大学', '西南交通大学', '电子科技大学',
    '四川农业大学', '西南财经大学', '贵州大学', '云南大学',
    '西藏大学', '西北大学', '西安交通大学', '西北工业大学',
    '西安电子科技大学', '长安大学', '西北农林科技大学', '陕西师范大学',
    '兰州大学', '青海大学', '宁夏大学', '新疆大学',
    '石河子大学', '湘潭大学', '山西大学', '南京医科大学',
    '广州医科大学', '华南农业大学', '南方科技大学', '上海科技大学',
    '中国科学院大学', '河南大学'
]

SCHOOL_INTERNATIONAL = [
    '西交利物浦大学', '宁波诺丁汉大学', '北京师范大学-香港浸会大学联合国际学院',
    '北师香港浸会大学', '上海纽约大学', '昆山杜克大学', '香港中文大学（深圳）',
    '深圳北理莫斯科大学', '广东以色列理工学院', '香港科技大学（广州）',
    '香港大学', '香港科技大学', '香港中文大学', '香港城市大学',
    '香港理工大学', '香港浸会大学', '香港岭南大学', '香港教育大学',
    '澳门科技大学', '澳门城市大学', '汕头大学'
]

header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
sub_header_fill = PatternFill(start_color='8FAADC', end_color='8FAADC', fill_type='solid')
title_font = Font(name='微软雅黑', size=14, bold=True, color='4472C4')
normal_font = Font(name='微软雅黑', size=10)
highlight_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

def style_header(ws, row=1, fill=None):
    if fill is None:
        fill = header_fill
    for cell in ws[row]:
        cell.font = header_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def style_data(ws, start_row=2):
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            cell.font = normal_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

def auto_width(ws, min_width=8, max_width=30):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                try:
                    cell_len = len(str(cell.value))
                    if cell_len > max_len:
                        max_len = cell_len
                except:
                    pass
        width = min(max(max_len * 1.2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width

def freeze_first_row(ws):
    ws.freeze_panes = 'A2'

def add_title(ws, title, cols):
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = title_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

def main():
    print('读取原始数据...')
    df = pd.read_excel(INPUT_FILE, sheet_name=0)
    print(f'共 {len(df)} 条数据，{df["school_name"].nunique()} 所院校')

    df = df.fillna('')

    df['is_985'] = df['school_name'].isin(SCHOOL_985)
    df['is_211'] = df['school_name'].isin(SCHOOL_211)
    df['is_double_first'] = df['school_name'].isin(SCHOOL_DOUBLE_FIRST)
    df['is_international'] = df['school_name'].isin(SCHOOL_INTERNATIONAL)

    def get_score_level(score):
        if score >= 800:
            return '高分段(800+)'
        elif score >= 700:
            return '中高分段(700-799)'
        elif score >= 600:
            return '中分段(600-699)'
        elif score >= 550:
            return '中低分段(550-599)'
        else:
            return '低分段(<550)'
    
    df['score_level'] = df['min_score'].apply(get_score_level)

    df_2025 = df[df['year'] == 2025].copy()
    df_2024 = df[df['year'] == 2024].copy()
    df_2023 = df[df['year'] == 2023].copy()

    print('生成Excel文件...')
    writer = pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl')

    sheet1_cols = ['学校名称', '学校层次', '2025最低分', '2025最低位次', '2025招生专业数',
                   '2024最低分', '2024最低位次', '2024招生专业数',
                   '2023最低分', '2023最低位次', '2023招生专业数',
                   '三年平均分变化', '选科要求种类']
    
    school_summary = []
    for school in sorted(df['school_name'].unique()):
        s_df = df[df['school_name'] == school]
        
        levels = []
        if school in SCHOOL_985:
            levels.append('985')
        if school in SCHOOL_211:
            levels.append('211')
        if school in SCHOOL_DOUBLE_FIRST:
            levels.append('双一流')
        if school in SCHOOL_INTERNATIONAL:
            levels.append('中外合作')
        level_str = '/'.join(levels) if levels else '普通本科'
        
        row_data = {'学校名称': school, '学校层次': level_str}
        
        for year in [2025, 2024, 2023]:
            y_df = s_df[s_df['year'] == year]
            if len(y_df) > 0:
                row_data[f'{year}最低分'] = y_df['min_score'].min()
                row_data[f'{year}最低位次'] = y_df['min_rank'].min()
                row_data[f'{year}招生专业数'] = y_df['major_name'].nunique()
            else:
                row_data[f'{year}最低分'] = ''
                row_data[f'{year}最低位次'] = ''
                row_data[f'{year}招生专业数'] = 0
        
        scores = []
        for year in [2023, 2024, 2025]:
            y_df = s_df[s_df['year'] == year]
            if len(y_df) > 0:
                scores.append(y_df['min_score'].min())
        if len(scores) >= 2:
            trend = scores[-1] - scores[0]
            if trend > 0:
                row_data['三年平均分变化'] = f'上升{trend}分'
            elif trend < 0:
                row_data['三年平均分变化'] = f'下降{abs(trend)}分'
            else:
                row_data['三年平均分变化'] = '持平'
        else:
            row_data['三年平均分变化'] = '-'
        
        reqs = s_df['subject_requirement'].unique()
        reqs = [r for r in reqs if r and r != '']
        row_data['选科要求种类'] = len(reqs)
        
        school_summary.append(row_data)
    
    df_summary = pd.DataFrame(school_summary)
    df_summary['2025最低分_num'] = pd.to_numeric(df_summary['2025最低分'], errors='coerce')
    df_summary = df_summary.sort_values('2025最低分_num', ascending=False, na_position='last')
    df_summary = df_summary.drop(columns=['2025最低分_num'])
    
    print('  Sheet1: 院校总览')
    df_summary.to_excel(writer, sheet_name='院校总览', index=False)

    detail_cols = ['学校名称', '年份', '专业名称', '专业组', '最低分', '最低位次',
                   '录取人数', '选科要求', '专业描述']

    def write_detail_sheet(df_data, sheet_name, title):
        d = df_data[['school_name', 'year', 'major_name', 'major_group', 'min_score',
                     'min_rank', 'person_count', 'subject_requirement', 'major_description']].copy()
        d.columns = detail_cols
        d = d.sort_values(['最低分', '最低位次'], ascending=[False, True])
        d.to_excel(writer, sheet_name=sheet_name, index=False)

    print('  Sheet2: 985院校')
    df_985 = df[df['is_985'] == True]
    write_detail_sheet(df_985, '985院校', '985院校专业分数线')

    print('  Sheet3: 211院校')
    df_211 = df[df['is_211'] == True]
    write_detail_sheet(df_211, '211院校', '211院校专业分数线')

    print('  Sheet4: 双一流院校')
    df_df = df[df['is_double_first'] == True]
    write_detail_sheet(df_df, '双一流院校', '双一流院校专业分数线')

    print('  Sheet5: 中外合作办学院校')
    df_intl = df[df['is_international'] == True]
    write_detail_sheet(df_intl, '中外合作办学院校', '中外合作办学院校专业分数线')

    print('  Sheet6: 按分数段分类(2025年)')
    score_level_data = []
    for level in ['高分段(800+)', '中高分段(700-799)', '中分段(600-699)', 
                  '中低分段(550-599)', '低分段(<550)']:
        level_df = df_2025[df_2025['score_level'] == level]
        if len(level_df) > 0:
            d = level_df[['school_name', 'major_name', 'major_group', 'min_score',
                         'min_rank', 'person_count', 'subject_requirement']].copy()
            d.columns = ['学校名称', '专业名称', '专业组', '最低分', '最低位次', '录取人数', '选科要求']
            d.insert(0, '分数段', level)
            d = d.sort_values('最低分', ascending=False)
            score_level_data.append(d)
    if score_level_data:
        df_score_level = pd.concat(score_level_data, ignore_index=True)
        df_score_level.to_excel(writer, sheet_name='按分数段分类', index=False)

    print('  Sheet7: 按选科要求分类(2025年)')
    subject_data = []
    for req in ['物+化(2科必选)', '不限', '必选物理', '必选政治', '必选历史',
                '物+化+生(3科必选)', '物/化/生(3选1)', '必选地理']:
        req_df = df_2025[df_2025['subject_requirement'] == req]
        if len(req_df) > 0:
            d = req_df[['school_name', 'major_name', 'major_group', 'min_score',
                       'min_rank', 'person_count']].copy()
            d.columns = ['学校名称', '专业名称', '专业组', '最低分', '最低位次', '录取人数']
            d.insert(0, '选科要求', req)
            d = d.sort_values('最低分', ascending=False)
            subject_data.append(d)
    if subject_data:
        df_subject = pd.concat(subject_data, ignore_index=True)
        df_subject.to_excel(writer, sheet_name='按选科要求分类', index=False)

    print('  Sheet8: 三年分数对比')
    comparison_data = []
    for school in sorted(df['school_name'].unique()):
        s_df = df[df['school_name'] == school]
        majors = s_df['major_name'].unique()
        for major in majors:
            m_df = s_df[s_df['major_name'] == major]
            groups = m_df['major_group'].unique()
            for group in groups:
                g_df = m_df[m_df['major_group'] == group]
                row = {
                    '学校名称': school,
                    '专业名称': major,
                    '专业组': group,
                }
                for year in [2023, 2024, 2025]:
                    y_df = g_df[g_df['year'] == year]
                    if len(y_df) > 0:
                        row[f'{year}最低分'] = y_df['min_score'].iloc[0]
                        row[f'{year}最低位次'] = y_df['min_rank'].iloc[0]
                    else:
                        row[f'{year}最低分'] = ''
                        row[f'{year}最低位次'] = ''
                
                s23 = row.get('2023最低分', '')
                s25 = row.get('2025最低分', '')
                if s23 and s25 and s23 != '' and s25 != '':
                    try:
                        diff = int(s25) - int(s23)
                        if diff > 0:
                            row['2023-2025分差'] = f'+{diff}'
                        elif diff < 0:
                            row['2023-2025分差'] = str(diff)
                        else:
                            row['2023-2025分差'] = '0'
                    except:
                        row['2023-2025分差'] = ''
                else:
                    row['2023-2025分差'] = ''
                
                comparison_data.append(row)
    
    df_comp = pd.DataFrame(comparison_data)
    df_comp['sort_score'] = pd.to_numeric(df_comp['2025最低分'], errors='coerce')
    df_comp = df_comp.sort_values('sort_score', ascending=False, na_position='last')
    df_comp = df_comp.drop(columns=['sort_score'])
    df_comp.to_excel(writer, sheet_name='三年分数对比', index=False)

    print('  Sheet9: 院校录取位次排名(2025)')
    rank_2025 = []
    for school in sorted(df_2025['school_name'].unique()):
        s_df = df_2025[df_2025['school_name'] == school]
        if len(s_df) > 0:
            levels = []
            if school in SCHOOL_985:
                levels.append('985')
            if school in SCHOOL_211:
                levels.append('211')
            if school in SCHOOL_DOUBLE_FIRST:
                levels.append('双一流')
            level_str = '/'.join(levels) if levels else '普通本科'
            
            rank_2025.append({
                '排名': 0,
                '学校名称': school,
                '学校层次': level_str,
                '最低分': s_df['min_score'].min(),
                '最低位次': s_df['min_rank'].min(),
                '最高分': s_df['min_score'].max(),
                '平均分': round(s_df['min_score'].mean(), 1),
                '招生专业数': s_df['major_name'].nunique(),
            })
    
    df_rank = pd.DataFrame(rank_2025)
    df_rank = df_rank.sort_values('最低分', ascending=False).reset_index(drop=True)
    df_rank['排名'] = range(1, len(df_rank) + 1)
    df_rank.to_excel(writer, sheet_name='院校位次排名', index=False)

    print('  Sheet10: 原始数据')
    df_orig = df[['school_name', 'year', 'major_name', 'major_group', 'min_score',
                  'min_rank', 'person_count', 'batch', 'subject_requirement',
                  'major_description', 'province']].copy()
    df_orig.columns = ['学校名称', '年份', '专业名称', '专业组', '最低分', '最低位次',
                       '录取人数', '批次', '选科要求', '专业描述', '省份']
    df_orig.to_excel(writer, sheet_name='原始数据', index=False)

    writer.close()

    print('美化排版...')
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    
    sheet_titles = {
        '院校总览': '海南高考志愿填报参考 - 院校总览',
        '985院校': '985院校专业分数线（海南）',
        '211院校': '211院校专业分数线（海南）',
        '双一流院校': '双一流院校专业分数线（海南）',
        '中外合作办学院校': '中外合作办学院校专业分数线（海南）',
        '按分数段分类': '2025年海南高考 - 按分数段分类',
        '按选科要求分类': '2025年海南高考 - 按选科要求分类',
        '三年分数对比': '海南高考专业分数线 - 三年对比',
        '院校位次排名': '2025年海南高考院校录取排名',
        '原始数据': '原始数据（全部记录）',
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f'  美化: {sheet_name}')
        
        if sheet_name in sheet_titles:
            add_title(ws, sheet_titles[sheet_name], ws.max_column)
            style_header(ws, row=2)
            style_data(ws, start_row=3)
            ws.freeze_panes = 'A3'
        else:
            style_header(ws)
            style_data(ws)
            freeze_first_row(ws)
        
        auto_width(ws)
    
    wb.save(OUTPUT_FILE)
    print(f'\n✅ 文件已生成: {OUTPUT_FILE}')
    print(f'   共 {len(wb.sheetnames)} 个工作表')

if __name__ == '__main__':
    main()
