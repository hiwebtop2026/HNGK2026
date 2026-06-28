# -*- coding: utf-8 -*-
"""
夸克高考专业分数线自动抓取工具 v3
修复：切换到专业分数线Tab、正确提取DOM数据
"""

import json
import time
import os
import re
from datetime import datetime

from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.microsoft import EdgeChromiumDriverManager

import openpyxl

EXCEL_FILE = r'C:\Users\lhp\Desktop\2023-2025年海南高考本科投档分数线.xlsx'
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data')

YEARS = [2025, 2024, 2023]
PROVINCE = '海南'
BATCH = '本科批'


def ensure_dir(dir_path):
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)


def read_schools_from_excel():
    """从Excel读取院校列表"""
    print('正在读取Excel文件获取院校列表...')
    schools = set()

    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)

        for year in ['2025', '2024', '2023']:
            if year not in wb.sheetnames:
                continue

            ws = wb[year]

            name_col_idx = None
            header_row_idx = 2

            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
                for col_idx, cell in enumerate(row):
                    if cell and isinstance(cell, str):
                        cell_str = cell.strip()
                        if ('院校' in cell_str and '名称' in cell_str) or cell_str == '院校专业组名称':
                            name_col_idx = col_idx
                            header_row_idx = row_idx
                            break
                if name_col_idx is not None:
                    break

            if name_col_idx is None:
                continue

            count = 0
            for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                if len(row) <= name_col_idx:
                    continue

                cell_val = row[name_col_idx]
                if not cell_val or not isinstance(cell_val, str):
                    continue

                cell_val = cell_val.strip()
                if len(cell_val) < 2:
                    continue

                if '科目' in cell_val and '要求' in cell_val:
                    continue
                if cell_val in ['说明', '注', '备注']:
                    continue

                school_name = re.sub(r'\([^)]*\)', '', cell_val).strip()
                school_name = re.sub(r'\d+$', '', school_name).strip()

                if len(school_name) >= 2 and not any(c.isdigit() for c in school_name[:2]):
                    schools.add(school_name)
                    count += 1

            print(f'  {year}年: {count}个')

        wb.close()
    except Exception as e:
        print(f'读取Excel失败: {e}')

    result = sorted(list(schools))
    print(f'共提取 {len(result)} 个唯一院校')
    return result


def setup_driver():
    """设置Edge浏览器"""
    print('正在启动Edge浏览器...')

    options = Options()
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_argument('--start-maximized')
    options.add_argument('--disable-infobars')
    options.add_argument('--disable-gpu')
    options.add_argument('--disable-dev-shm-usage')
    options.add_experimental_option('excludeSwitches', ['enable-automation'])
    options.add_experimental_option('useAutomationExtension', False)
    options.page_load_timeout = 60000

    try:
        driver = webdriver.Edge(options=options)
    except Exception:
        print('正在下载Edge驱动...')
        service = Service(EdgeChromiumDriverManager().install())
        driver = webdriver.Edge(service=service, options=options)

    driver.set_page_load_timeout(60)

    driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
        'source': '''
            Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
            window.chrome = { runtime: {} };
        '''
    })

    return driver


def build_school_url(school_name, year):
    """构建院校页面URL"""
    params = json.dumps({
        'province': PROVINCE,
        'year': str(year),
        'batch': BATCH,
        'genre': '综合'
    }, ensure_ascii=False)

    return (f'https://vt.quark.cn/blm/gaokao-college-794/tab?app=fen_shu_xian'
            f'&university_name={school_name}'
            f'&q={school_name}'
            f'&device=pc'
            f'&params={params}'
            f'&type=zhuanye')


def switch_to_major_tab(driver):
    """切换到专业分数线Tab"""
    try:
        script = '''
        const tabs = document.querySelectorAll('.qk-tabs-tab, [class*="tab"]');
        for (const tab of tabs) {
            const text = tab.textContent || '';
            if (text.includes('专业分数线') || text.includes('专业')) {
                tab.click();
                return { clicked: true, text: text.trim() };
            }
        }
        return { clicked: false };
        '''
        result = driver.execute_script(script)
        if result and result.get('clicked'):
            print(f'  已切换到: {result.get("text")}')
            time.sleep(3)
            return True
    except Exception as e:
        print(f'  切换Tab失败: {e}')

    return False


def click_view_all(driver):
    """点击查看全部"""
    try:
        script = '''
        const moreBtns = document.querySelectorAll('.qk-more, [class*="more"]');
        for (const btn of moreBtns) {
            const text = btn.textContent || '';
            if (text.includes('查看全部') || text.includes('全部')) {
                btn.click();
                return { clicked: true, text: text.trim() };
            }
        }

        const allBtns = document.querySelectorAll('button, a, div[role="button"]');
        for (const btn of allBtns) {
            const text = btn.textContent || '';
            if (text.trim() === '查看全部' || text.trim() === '全部') {
                btn.click();
                return { clicked: true, text: text.trim() };
            }
        }

        return { clicked: false };
        '''
        result = driver.execute_script(script)
        if result and result.get('clicked'):
            print(f'  点击了: {result.get("text")}')
            time.sleep(3)
            return True
    except Exception as e:
        print(f'  点击查看全部失败: {e}')

    return False


def switch_year(driver, year):
    """切换年份"""
    try:
        year_str = str(year)
        script = f'''
        const yearBtns = document.querySelectorAll('.select-tabs-tab-nianfen, [class*="year"], [class*="nianfen"]');
        for (const btn of yearBtns) {{
            const text = btn.textContent || '';
            if (text.includes('{year_str}')) {{
                btn.click();
                return {{ switched: true, text: text.trim() }};
            }}
        }}

        const allBtns = document.querySelectorAll('button');
        for (const btn of allBtns) {{
            const text = (btn.textContent || '').trim();
            if (text === '{year_str}' || text.startsWith('{year_str}')) {{
                btn.click();
                return {{ switched: true, text: text }};
            }}
        }}

        return {{ switched: false }};
        '''
        result = driver.execute_script(script)
        if result and result.get('switched'):
            print(f'  已切换年份: {result.get("text")}')
            time.sleep(3)
            return True
    except Exception as e:
        print(f'  切换年份失败: {e}')

    return False


def switch_province(driver, province):
    """切换省份"""
    try:
        script = f'''
        const cityBtns = document.querySelectorAll('.select-tabs-tab-chengshi, [class*="city"], [class*="chengshi"]');
        for (const btn of cityBtns) {{
            const text = btn.textContent || '';
            if (text.includes('{province}')) {{
                btn.click();
                return {{ switched: true, text: text.trim() }};
            }}
        }}

        return {{ switched: false }};
        '''
        result = driver.execute_script(script)
        if result and result.get('switched'):
            print(f'  已切换省份: {result.get("text")}')
            time.sleep(2)
            return True
    except Exception as e:
        print(f'  切换省份失败: {e}')

    return False


def extract_major_scores_js(driver, school_name, year):
    """通过JavaScript提取专业分数线数据"""
    script = '''
    function extractMajorScores(schoolName, year, province, batch) {
        const results = [];

        // 方法1: 从专业分数线列表提取
        const listItems = document.querySelectorAll('.content-List-li, [class*="content-List-li"]');

        listItems.forEach(item => {
            // 跳过表头
            if (item.querySelector('.content-icon')) return;

            const majorEl = item.querySelector('[class*="content-List-major"]');
            const scoreEl = item.querySelector('[class*="content-List-low_score"]');
            const rankEl = item.querySelector('[class*="content-List-low_rank"]');
            const countEl = item.querySelector('[class*="content-List-luqurenshu"]');
            const diffEl = item.querySelector('[class*="content-List-low_score_diff"]');

            const majorName = majorEl ? (majorEl.textContent || '').trim() : '';
            const scoreText = scoreEl ? (scoreEl.textContent || '').trim() : '';
            const rankText = rankEl ? (rankEl.textContent || '').trim() : '';
            const countText = countEl ? (countEl.textContent || '').trim() : '';
            const diffText = diffEl ? (diffEl.textContent || '').trim() : '';

            if (!majorName && !scoreText) return;

            const record = {
                school_name: schoolName,
                province: province,
                year: year,
                batch: batch,
                source: '夸克高考',
            };

            if (majorName && majorName !== '--') {
                record.major_name = majorName;
            }

            if (scoreText && scoreText !== '--') {
                const score = parseInt(scoreText);
                if (!isNaN(score)) record.min_score = score;
            }

            if (rankText && rankText !== '--') {
                const rank = parseInt(rankText.replace(/[^\\d]/g, ''));
                if (!isNaN(rank)) record.min_rank = rank;
            }

            if (countText && countText !== '--') {
                const count = parseInt(countText);
                if (!isNaN(count)) record.person_count = count;
            }

            if (diffText && diffText !== '--') {
                const diff = parseInt(diffText);
                if (!isNaN(diff)) record.batch_line_diff = diff;
            }

            // 查找专业组和科目要求（在副标题中）
            let parent = item.parentElement;
            let subTitle = '';
            let batchText = '';
            for (let i = 0; i < 5 && parent; i++) {
                const subTitleEl = parent.querySelector('[class*="content-List-subTitle"]');
                const batchEl = parent.querySelector('[class*="qk-paragraph"] [class*="qk-color-dark"]');
                if (subTitleEl && !subTitle) {
                    subTitle = (subTitleEl.textContent || '').trim();
                }
                if (batchEl && !batchText) {
                    batchText = (batchEl.textContent || '').trim();
                    if (batchText && !batchText.includes('批')) {
                        batchText = '';
                    }
                }
                parent = parent.parentElement;
            }

            if (subTitle) {
                const groupMatch = subTitle.match(/专业组([^-]+)/);
                if (groupMatch) {
                    record.major_group = '专业组' + groupMatch[1].trim();
                }
                const subjectMatch = subTitle.match(/[-–]\s*(.+)/);
                if (subjectMatch) {
                    record.subject_requirement = subjectMatch[1].trim();
                }
            }

            if (batchText) {
                record.batch = batchText;
            }

            if (record.major_name || record.min_score) {
                results.push(record);
            }
        });

        // 方法2: 如果没有找到，尝试从所有列表项提取
        if (results.length === 0) {
            const allItems = document.querySelectorAll('[class*="list-item"], [class*="List-li"], li');
            allItems.forEach(item => {
                const text = item.textContent || '';
                if (text.length < 5 || text.length > 200) return;

                const scoreMatch = text.match(/(\d{2,3})分/);
                const rankMatch = text.match(/(\d{3,6})位次/);
                const majorMatch = text.match(/^([^\s\d]{2,}(?:专业|学|工程|技术|管理|经济|法学|文学|理学|工学|医学|农学|军事|艺术|教育|历史|哲学)[^\s]*)/m);

                if (scoreMatch) {
                    const record = {
                        school_name: schoolName,
                        major_name: majorMatch ? majorMatch[1] : text.substring(0, 30),
                        min_score: parseInt(scoreMatch[1]),
                        province: province,
                        year: year,
                        batch: batch,
                        source: '夸克高考',
                    };

                    if (rankMatch) {
                        record.min_rank = parseInt(rankMatch[1]);
                    }

                    results.push(record);
                }
            });
        }

        return results;
    }

    return extractMajorScores(arguments[0], arguments[1], arguments[2], arguments[3]);
    '''

    try:
        results = driver.execute_script(script, school_name, year, PROVINCE, BATCH)
        return results if results else []
    except Exception as e:
        print(f'  JS提取出错: {e}')
        return []


def main():
    print('=' * 70)
    print('夸克高考专业分数线自动抓取工具 v3')
    print('=' * 70)

    ensure_dir(OUTPUT_DIR)

    schools = read_schools_from_excel()
    if not schools:
        print('未找到院校数据，退出')
        return

    driver = setup_driver()

    all_records = []
    success_count = 0
    fail_count = 0

    try:
        test_school = schools[0]
        print(f'\n测试院校: {test_school}')

        for year in YEARS:
            print(f'\n  === {year}年 ===')

            url = build_school_url(test_school, year)
            print(f'  访问页面...')

            try:
                driver.get(url)
            except Exception as e:
                print(f'  页面加载超时: {e}')

            print('  等待页面加载...')
            time.sleep(8)

            print('  切换到专业分数线Tab...')
            switch_to_major_tab(driver)

            print('  点击查看全部...')
            click_view_all(driver)

            time.sleep(3)

            print('  提取专业分数线数据...')
            records = extract_major_scores_js(driver, test_school, year)
            print(f'  提取到 {len(records)} 条数据')

            if records:
                all_records.extend(records)
                success_count += 1

                year_file = os.path.join(OUTPUT_DIR, f'major_scores_{year}.json')
                with open(year_file, 'w', encoding='utf-8') as f:
                    json.dump(records, f, ensure_ascii=False, indent=2)
                print(f'  已保存到: {year_file}')

                print('  数据预览:')
                for i, r in enumerate(records[:3]):
                    print(f'    {i+1}. {r.get("major_name")} - {r.get("min_score")}分 - {r.get("min_rank")}位次')
            else:
                fail_count += 1

            screenshot_file = os.path.join(OUTPUT_DIR, f'screenshot_{year}.png')
            try:
                driver.save_screenshot(screenshot_file)
            except Exception:
                pass

            time.sleep(2)

        all_file = os.path.join(OUTPUT_DIR, 'major_scores_all.json')
        with open(all_file, 'w', encoding='utf-8') as f:
            json.dump(all_records, f, ensure_ascii=False, indent=2)

        print('\n' + '=' * 70)
        print('测试抓取完成！')
        print(f'共获取 {len(all_records)} 条专业分数线数据')
        print(f'成功: {success_count} 年 | 失败: {fail_count} 年')
        print(f'数据已保存到: {all_file}')
        print('=' * 70)

        print('\n浏览器将在60秒后自动关闭...')
        time.sleep(60)

    except Exception as e:
        print(f'抓取出错: {e}')
        import traceback
        traceback.print_exc()
        time.sleep(30)
    finally:
        try:
            driver.quit()
        except Exception:
            pass
        print('浏览器已关闭')


if __name__ == '__main__':
    main()
