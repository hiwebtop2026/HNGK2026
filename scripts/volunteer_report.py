import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER

def extract_school_name(school_name):
    if pd.isna(school_name):
        return school_name
    name = str(school_name)
    import re
    match = re.match(r'^(.+?)(?:\(\d+\))?$', name)
    return match.group(1).strip() if match else name.strip()

def is_subject_match(subject_requirement):
    if pd.isna(subject_requirement):
        return True
    req = str(subject_requirement).strip()
    if req == '不限':
        return True
    if '物' in req:
        return True
    return False

def calculate_trend(row):
    scores = []
    for year in [2023, 2024, 2025]:
        score = row.get(f'{year}分数线')
        if pd.notna(score) and isinstance(score, (int, float)):
            scores.append((year, score))
    
    if len(scores) < 2:
        return '数据不足'
    
    recent = scores[-1]
    previous = scores[-2]
    
    diff = recent[1] - previous[1]
    if diff > 5:
        return '上涨'
    elif diff < -5:
        return '下降'
    else:
        return '稳定'

def set_cell_style(ws, row, col, value, fill=None, font=None, alignment=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if number_format:
        cell.number_format = number_format

def main():
    input_file = r'G:\Downloads\20260703_志愿匹配.xlsx'
    output_file = r'G:\Downloads\20260703_志愿填报分析报告.xlsx'
    
    student_score = 603
    student_rank = 12085
    student_subjects = '物化生'
    
    print(f"处理文件: {input_file}")
    print(f"考生信息: 分数={student_score}, 位次={student_rank}, 选科={student_subjects}")
    
    try:
        df = pd.read_excel(input_file)
        print(f"成功读取文件，共 {len(df)} 行数据")
        
        filtered_df = df[df['志愿匹配'].notna()].copy()
        filtered_df = filtered_df[filtered_df.apply(lambda row: is_subject_match(row['subject_requirement']), axis=1)]
        print(f"选科匹配后有效数据: {len(filtered_df)} 行")
        
        filtered_df['院校主体'] = filtered_df['school_name'].apply(extract_school_name)
        
        chong_df = filtered_df[filtered_df['志愿匹配'] == '冲']
        wen_df = filtered_df[filtered_df['志愿匹配'] == '稳']
        bao_df = filtered_df[filtered_df['志愿匹配'] == '保']
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            pd.DataFrame({
                '指标': ['考生分数', '考位次', '选科组合', '总数据行数', '有效匹配行数', '冲刺志愿数', '稳妥志愿数', '保底志愿数'],
                '数值': [student_score, student_rank, student_subjects, len(df), len(filtered_df), len(chong_df), len(wen_df), len(bao_df)]
            }).to_excel(writer, sheet_name='考生信息', index=False)
            
            school_stats = filtered_df.groupby(['院校主体', '志愿匹配']).agg({
                'min_score': ['min', 'max', 'mean'],
                'min_rank': ['min', 'max', 'mean'],
                'major_name': 'count',
                'year': 'nunique',
                'school_name': 'nunique'
            }).reset_index()
            school_stats.columns = ['院校名称', '志愿匹配', '最低分', '最高分', '平均分', '最低位次', '最高位次', '平均位次', '专业数', '年份数', '专业组数']
            school_stats.to_excel(writer, sheet_name='院校统计', index=False)
            
            major_stats = filtered_df.groupby(['major_name', '志愿匹配']).agg({
                'min_score': ['min', 'max', 'mean'],
                'min_rank': ['min', 'max', 'mean'],
                '院校主体': 'nunique',
                'year': 'nunique'
            }).reset_index()
            major_stats.columns = ['专业名称', '志愿匹配', '最低分', '最高分', '平均分', '最低位次', '最高位次', '平均位次', '院校数', '年份数']
            major_stats.to_excel(writer, sheet_name='专业统计', index=False)
            
            yearly_stats = filtered_df.groupby(['year', '志愿匹配']).agg({
                'min_score': ['min', 'max', 'mean'],
                'min_rank': ['min', 'max', 'mean'],
                '院校主体': 'nunique',
                'major_name': 'count'
            }).reset_index()
            yearly_stats.columns = ['年份', '志愿匹配', '最低分', '最高分', '平均分', '最低位次', '最高位次', '平均位次', '院校数', '专业数']
            yearly_stats.to_excel(writer, sheet_name='年份统计', index=False)
            
            chong_df.to_excel(writer, sheet_name='冲刺志愿详情', index=False)
            wen_df.to_excel(writer, sheet_name='稳妥志愿详情', index=False)
            bao_df.to_excel(writer, sheet_name='保底志愿详情', index=False)
            
            pivot_year = filtered_df.pivot_table(
                index=['院校主体', 'major_name'],
                columns='year',
                values='min_score',
                aggfunc='min'
            ).reset_index()
            pivot_year.columns = ['院校名称', '专业名称', '2023分数线', '2024分数线', '2025分数线']
            pivot_year['分数线趋势'] = pivot_year.apply(calculate_trend, axis=1)
            pivot_year['平均分数线'] = pivot_year[['2023分数线', '2024分数线', '2025分数线']].mean(axis=1)
            pivot_year['分数差'] = student_score - pivot_year['平均分数线']
            pivot_year.to_excel(writer, sheet_name='历年分数线对比', index=False)
            
            tier_dist = filtered_df.groupby('院校主体')['志愿匹配'].value_counts().unstack(fill_value=0)
            tier_dist = tier_dist[['冲', '稳', '保']].fillna(0)
            tier_dist['总计'] = tier_dist.sum(axis=1)
            tier_dist['占比(冲)'] = (tier_dist['冲'] / tier_dist['总计'] * 100).round(1)
            tier_dist['占比(稳)'] = (tier_dist['稳'] / tier_dist['总计'] * 100).round(1)
            tier_dist['占比(保)'] = (tier_dist['保'] / tier_dist['总计'] * 100).round(1)
            tier_dist.to_excel(writer, sheet_name='院校冲稳保分布')
        
        wb = openpyxl.load_workbook(output_file)
        
        yellow_fill = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid')
        green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
        blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
        gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        center_align = Alignment(horizontal='center', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    if row == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center_align
            
            if sheet_name == '考生信息':
                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 15
                
            elif sheet_name == '院校统计':
                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 10
                for col in range(3, 14):
                    ws.column_dimensions[get_column_letter(col)].width = 12
                
                for row in range(2, ws.max_row + 1):
                    tier_cell = ws.cell(row=row, column=2)
                    if tier_cell.value == '保':
                        tier_cell.fill = green_fill
                    elif tier_cell.value == '稳':
                        tier_cell.fill = yellow_fill
                    elif tier_cell.value == '冲':
                        tier_cell.fill = red_fill
            
            elif sheet_name == '专业统计':
                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 10
                for col in range(3, 12):
                    ws.column_dimensions[get_column_letter(col)].width = 12
                
                for row in range(2, ws.max_row + 1):
                    tier_cell = ws.cell(row=row, column=2)
                    if tier_cell.value == '保':
                        tier_cell.fill = green_fill
                    elif tier_cell.value == '稳':
                        tier_cell.fill = yellow_fill
                    elif tier_cell.value == '冲':
                        tier_cell.fill = red_fill
            
            elif sheet_name == '年份统计':
                ws.column_dimensions['A'].width = 10
                ws.column_dimensions['B'].width = 10
                for col in range(3, 12):
                    ws.column_dimensions[get_column_letter(col)].width = 12
                
                for row in range(2, ws.max_row + 1):
                    tier_cell = ws.cell(row=row, column=2)
                    if tier_cell.value == '保':
                        tier_cell.fill = green_fill
                    elif tier_cell.value == '稳':
                        tier_cell.fill = yellow_fill
                    elif tier_cell.value == '冲':
                        tier_cell.fill = red_fill
            
            elif sheet_name in ['冲刺志愿详情', '稳妥志愿详情', '保底志愿详情']:
                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 10
                ws.column_dimensions['C'].width = 20
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 12
                ws.column_dimensions['F'].width = 12
                ws.column_dimensions['G'].width = 10
                ws.column_dimensions['H'].width = 10
                ws.column_dimensions['I'].width = 30
                ws.column_dimensions['J'].width = 20
                ws.column_dimensions['K'].width = 10
                ws.column_dimensions['L'].width = 12
                ws.column_dimensions['M'].width = 20
                
                for row in range(2, ws.max_row + 1):
                    tier_cell = ws.cell(row=row, column=12)
                    if sheet_name == '冲刺志愿详情':
                        tier_cell.fill = red_fill
                        tier_cell.font = Font(color='FFFFFF', bold=True)
                    elif sheet_name == '稳妥志愿详情':
                        tier_cell.fill = yellow_fill
                        tier_cell.font = Font(color='000000', bold=True)
                    elif sheet_name == '保底志愿详情':
                        tier_cell.fill = green_fill
                        tier_cell.font = Font(color='FFFFFF', bold=True)
            
            elif sheet_name == '历年分数线对比':
                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 20
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 15
                ws.column_dimensions['F'].width = 12
                ws.column_dimensions['G'].width = 12
                ws.column_dimensions['H'].width = 12
                
                for row in range(2, ws.max_row + 1):
                    trend_cell = ws.cell(row=row, column=6)
                    if trend_cell.value == '上涨':
                        trend_cell.fill = red_fill
                        trend_cell.font = Font(color='FFFFFF', bold=True)
                    elif trend_cell.value == '下降':
                        trend_cell.fill = green_fill
                        trend_cell.font = Font(color='FFFFFF', bold=True)
                    elif trend_cell.value == '稳定':
                        trend_cell.fill = blue_fill
                        trend_cell.font = Font(color='000000', bold=True)
                    elif trend_cell.value == '数据不足':
                        trend_cell.fill = gray_fill
                        trend_cell.font = Font(color='000000', bold=True)
            
            elif sheet_name == '院校冲稳保分布':
                ws.column_dimensions['A'].width = 20
                for col in range(2, 10):
                    ws.column_dimensions[get_column_letter(col)].width = 10
                
                for row in range(2, ws.max_row + 1):
                    for col in range(2, 6):
                        cell = ws.cell(row=row, column=col)
                        cell.number_format = FORMAT_NUMBER
                        if col == 2:
                            cell.fill = red_fill
                            cell.font = Font(color='FFFFFF', bold=True)
                        elif col == 3:
                            cell.fill = yellow_fill
                            cell.font = Font(color='000000', bold=True)
                        elif col == 4:
                            cell.fill = green_fill
                            cell.font = Font(color='FFFFFF', bold=True)
        
        wb.save(output_file)
        print(f"\n处理完成，已保存到: {output_file}")
        print(f"\n生成的工作表:")
        for i, sheet in enumerate(wb.sheetnames, 1):
            print(f"  {i}. {sheet}")
        
        print(f"\n统计摘要:")
        print(f"  冲刺志愿: {len(chong_df)} 个")
        print(f"  稳妥志愿: {len(wen_df)} 个")
        print(f"  保底志愿: {len(bao_df)} 个")
        print(f"  涉及院校数: {filtered_df['院校主体'].nunique()} 所")
        print(f"  涉及专业数: {filtered_df['major_name'].nunique()} 个")
        
    except Exception as e:
        print(f"处理过程中发生错误: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()